r"""
97_ratios.py — Financial ratio time series

Compute standard financial ratios across N fiscal years from XBRLS multi-period
data (same source as 98_multiperiod_financials.py) and export a styled Excel
sheet: ratios as rows, fiscal years as columns.

Ratio groups:
  Profitability: Gross Margin, Operating Margin, Net Margin, ROIC, ROE
  Leverage:      Net Debt / EBITDA, Interest Coverage
  Liquidity:     Current Ratio, DSO, DIO, DPO, Cash Conversion Cycle
  FCF & Value:   Free Cash Flow, FCF Margin, FCF Yield (current mkt cap via yfinance)

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\97_ratios.py
"""
import os
import re
import time
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
os.environ["EDGAR_IDENTITY"] = os.getenv("SEC_USER_AGENT", "")

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from edgar import Company
from edgar.xbrl import XBRLS

# ── Configuration ─────────────────────────────────────────────────────────────
TICKER = "AAPL"
YEARS  = 10
SHOW_ITEM_MATCHES = False  # set True to print which IS/BS/CF row each lookup selects

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"
CLR_HDR_BG     = "4472C4"
CLR_SECTION_BG = "D9E1F2"
CLR_ALT_BG     = "F9F9F9"
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"
CLR_NA         = "999999"
THIN           = Side(style="thin",   color="CCCCCC")
MED            = Side(style="medium", color="AAAAAA")

FMT_PCT   = "0.0%"
FMT_X     = '0.0"x"'
FMT_DAYS  = "0"
FMT_USD_M = "#,##0"

# ── Ratio layout (defines Excel row order) ────────────────────────────────────
RATIO_SECTIONS = [
    {
        "title": "Profitability",
        "ratios": [
            {"label": "Gross Margin",      "key": "gross_margin",  "fmt": FMT_PCT},
            {"label": "Operating Margin",  "key": "op_margin",     "fmt": FMT_PCT},
            {"label": "Net Margin",        "key": "net_margin",    "fmt": FMT_PCT},
            {"label": "ROIC",              "key": "roic",          "fmt": FMT_PCT},
            {"label": "ROE",               "key": "roe",           "fmt": FMT_PCT},
        ],
    },
    {
        "title": "Leverage",
        "ratios": [
            {"label": "Net Debt / EBITDA", "key": "nd_ebitda",     "fmt": FMT_X},
            {"label": "Interest Coverage", "key": "int_coverage",  "fmt": FMT_X},
        ],
    },
    {
        "title": "Liquidity",
        "ratios": [
            {"label": "Current Ratio",       "key": "current_ratio", "fmt": FMT_X},
            {"label": "DSO (days)",          "key": "dso",           "fmt": FMT_DAYS},
            {"label": "DIO (days)",          "key": "dio",           "fmt": FMT_DAYS},
            {"label": "DPO (days)",          "key": "dpo",           "fmt": FMT_DAYS},
            {"label": "Cash Conv. Cycle",    "key": "ccc",           "fmt": FMT_DAYS},
        ],
    },
    {
        "title": "FCF & Valuation",
        "ratios": [
            {"label": "Free Cash Flow ($M)", "key": "fcf",          "fmt": FMT_USD_M},
            {"label": "FCF Margin",          "key": "fcf_margin",   "fmt": FMT_PCT},
            {"label": "FCF Yield *",         "key": "fcf_yield",    "fmt": FMT_PCT},
        ],
    },
]


# ── Utilities ─────────────────────────────────────────────────────────────────

def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


def _period_cols(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if re.match(r"\d{4}-\d{2}-\d{2}$", c)]


def _fy_label(date_str: str) -> str:
    return f"FY{date_str[:4]}"


# ── Data loading (same pattern as 98_multiperiod_financials.py) ───────────────

def _prepare_structured_df(
    template_stmt,
    stitched_stmt,
) -> tuple[pd.DataFrame, list[str]]:
    """Merge single-filing structure with XBRLS multi-period values.
    Returns (df, period_cols). Values scaled to USD millions."""
    tmpl     = template_stmt.to_dataframe(view="standard")
    stitched = stitched_stmt.to_dataframe()
    period_cols = sorted(_period_cols(stitched))

    for c in period_cols:
        stitched[c] = pd.to_numeric(stitched[c], errors="coerce")

    stitch_lookup: dict = {}
    for _, srow in stitched.iterrows():
        concept = srow["concept"]
        if concept not in stitch_lookup:
            stitch_lookup[concept] = {c: srow[c] for c in period_cols}

    rows = []
    for _, trow in tmpl.iterrows():
        if trow.get("dimension", False):
            continue
        concept     = trow["concept"]
        sc          = trow.get("standard_concept")
        is_abstract = bool(trow.get("abstract", False))
        row_data = {
            "label":            trow["label"],
            "level":            int(trow.get("level", 1)),
            "abstract":         is_abstract,
            "standard_concept": sc if pd.notna(sc) else "",
        }
        if is_abstract:
            for c in period_cols:
                row_data[c] = None
        else:
            hist = stitch_lookup.get(concept, {})
            for c in period_cols:
                row_data[c] = hist.get(c)
        rows.append(row_data)

    df = pd.DataFrame(rows)
    data_mask = ~df["abstract"]
    df = df[df["abstract"] | df.loc[data_mask, period_cols].notna().any(axis=1)].reset_index(drop=True)

    for idx, row in df.iterrows():
        if row["abstract"]:
            continue
        vals = [row[c] for c in period_cols if pd.notna(row[c])]
        if vals and max(abs(v) for v in vals) >= 1_000:
            for c in period_cols:
                if pd.notna(df.at[idx, c]):
                    df.at[idx, c] = df.at[idx, c] / 1_000_000

    return df, period_cols


# ── Line item lookup ──────────────────────────────────────────────────────────

def _get_item(
    df: pd.DataFrame,
    period_cols: list[str],
    std_concepts: list[str] | None = None,
    label_patterns: list[str] | None = None,
    exclude_patterns: list[str] | None = None,
    _label: str = "",
) -> dict | None:
    """Return {period_date: value} for first matching non-abstract row, or None.
    Tries standard_concept exact → label contains. Rows whose label contains any
    exclude_pattern are filtered out before matching."""
    data = df[~df["abstract"]]
    if exclude_patterns:
        excl = pd.Series(False, index=data.index)
        for pat in exclude_patterns:
            excl |= data["label"].str.lower().str.contains(pat.lower(), na=False)
        data = data[~excl]

    def _return(row, via):
        if _label and SHOW_ITEM_MATCHES:
            lv  = row.get("level", "?")
            sc  = row.get("standard_concept", "")
            lbl = row.get("label", "")
            print(f"  {_label:20s} [{via}] lv={lv}  label='{lbl}'  sc='{sc}'")
        return {c: pd.to_numeric(row.get(c), errors="coerce") for c in period_cols}

    def _best(matches):
        if "level" in matches.columns:
            m = matches.copy()
            m["_llen"] = m["label"].str.len()
            matches = m.sort_values(["level", "_llen"])
        return matches.iloc[0]

    if std_concepts:
        for sc in std_concepts:
            mask = data["standard_concept"].str.lower() == sc.lower()
            if mask.any():
                return _return(_best(data[mask]), "sc-exact")
        # sc-contains deliberately omitted: substring matching causes false positives
        # (e.g. "NonCurrentAssets" contains "CurrentAssets") — use exact match + labels only

    if label_patterns:
        for pat in label_patterns:
            mask = data["label"].str.lower().str.contains(pat.lower(), na=False)
            if mask.any():
                matches = data[mask]
                if "level" in matches.columns:
                    m = matches.copy()
                    m["_llen"] = m["label"].str.len()
                    matches = m.sort_values(["level", "_llen"])
                return _return(matches.iloc[0], f"label:'{pat}'")

    if _label and SHOW_ITEM_MATCHES:
        print(f"  {_label:20s} NOT FOUND")
    return None


# ── Arithmetic helpers ────────────────────────────────────────────────────────

def _ratio(
    num: dict | None,
    den: dict | None,
    period_cols: list[str],
    scale: float = 1.0,
) -> dict:
    """Element-wise num/den * scale. Missing or zero denominator -> None."""
    result = {}
    for c in period_cols:
        nv = (num or {}).get(c)
        dv = (den or {}).get(c)
        if pd.notna(nv) and pd.notna(dv) and dv != 0:
            result[c] = nv / dv * scale
        else:
            result[c] = None
    return result


def _add(
    a: dict | None,
    b: dict | None,
    period_cols: list[str],
    allow_partial: bool = False,
) -> dict:
    """Element-wise a + b. With allow_partial, treats missing as 0."""
    result = {}
    for c in period_cols:
        av = (a or {}).get(c)
        bv = (b or {}).get(c)
        av_ok = pd.notna(av)
        bv_ok = pd.notna(bv)
        if av_ok and bv_ok:
            result[c] = av + bv
        elif allow_partial and av_ok:
            result[c] = av
        elif allow_partial and bv_ok:
            result[c] = bv
        else:
            result[c] = None
    return result


def _avg_adjacent(series: dict, period_cols: list[str]) -> dict:
    """Average each period with its predecessor (for ROE)."""
    result = {}
    for i, c in enumerate(period_cols):
        curr = series.get(c)
        if i == 0 or not pd.notna(curr):
            result[c] = curr
        else:
            prev = series.get(period_cols[i - 1])
            result[c] = (curr + prev) / 2 if pd.notna(prev) else curr
    return result


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_ratio_sheet(
    wb: Workbook,
    ratios: dict,
    period_cols: list[str],
    fy_labels: list[str],
    company_name: str,
    form_type: str,
    mkt_cap_bn: float | None,
) -> None:
    n_cols = 1 + len(period_cols)
    ws = wb.create_sheet(title="Ratios")

    # ── Title ─────────────────────────────────────────────────────────────
    mkt_note = (
        f"  |  * FCF Yield uses current mkt cap: ${mkt_cap_bn:.1f}B"
        if mkt_cap_bn else "  |  * FCF Yield: market cap unavailable"
    )
    title_text = f"{company_name} ({TICKER})  -  Financial Ratios  ({form_type}){mkt_note}"
    ws.append([title_text] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1)
    tc.font      = Font(bold=True, color=CLR_WHITE, size=11)
    tc.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ── Column headers ────────────────────────────────────────────────────
    ws.append(["Ratio"] + fy_labels)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HDR_BG)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 1 else "left",
            vertical="center",
            indent=1,
        )
        cell.border = Border(bottom=Side(style="medium", color=CLR_WHITE))
    ws.row_dimensions[2].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    odd = True
    for section in RATIO_SECTIONS:
        ws.append([section["title"]] + [""] * (n_cols - 1))
        sec_row = ws.max_row
        ws.merge_cells(start_row=sec_row, start_column=1, end_row=sec_row, end_column=n_cols)
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(sec_row, col_idx)
            cell.fill      = PatternFill("solid", fgColor=CLR_SECTION_BG)
            cell.font      = Font(bold=True, size=10, color=CLR_DARK)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border    = Border(bottom=THIN)
        ws.row_dimensions[sec_row].height = 16
        odd = True

        for rdef in section["ratios"]:
            vals = ratios.get(rdef["key"], {})
            row_vals = [vals.get(c) for c in period_cols]
            ws.append([rdef["label"]] + row_vals)
            data_row = ws.max_row

            bg = PatternFill("solid", fgColor=CLR_ALT_BG) if odd else None
            odd = not odd

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(data_row, col_idx)
                if bg:
                    cell.fill = bg
                cell.border = Border(bottom=THIN)
                cell.font   = Font(size=10, color=CLR_DARK)

            ws.cell(data_row, 1).alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )

            for col_idx, c in enumerate(period_cols, 2):
                cell = ws.cell(data_row, col_idx)
                val  = vals.get(c)
                if val is not None and pd.notna(val):
                    cell.number_format = rdef["fmt"]
                else:
                    cell.value = "N/A"
                    cell.font  = Font(size=10, color=CLR_NA)
                cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[data_row].height = 15

    ws.column_dimensions["A"].width = 30
    for i in range(len(period_cols)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    ws.freeze_panes = "B3"


# ── Main ──────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

company = Company(TICKER)
step(f"Looked up {company.name} ({TICKER})  CIK {company.cik}", t_start)

form_type = None
filings   = None
for form in ("10-K", "20-F"):
    candidate = company.get_filings(form=form, amendments=False).head(YEARS)
    if len(candidate) > 0:
        form_type = form
        filings   = candidate
        break

if not filings:
    raise SystemExit(f"No 10-K or 20-F filings found for {TICKER}.")

step(
    f"Found {len(filings)} {form_type} filing(s) "
    f"({filings[len(filings)-1].filing_date} to {filings[0].filing_date})",
    t_start,
)

step("Getting latest filing structure (template)...", t_start)
template_financials = company.get_financials()
if template_financials is None:
    raise SystemExit("Could not parse financials from latest filing.")

step("Stitching XBRL across filings (may take 30-90s first run)...", t_start)
xbrls = XBRLS.from_filings(filings)
step("XBRLS ready", t_start)

step("Extracting statements...", t_start)
is_df, is_pc = _prepare_structured_df(
    template_financials.income_statement(),
    xbrls.statements.income_statement(),
)
bs_df, bs_pc = _prepare_structured_df(
    template_financials.balance_sheet(),
    xbrls.statements.balance_sheet(),
)
cf_df, cf_pc = _prepare_structured_df(
    template_financials.cashflow_statement(),
    xbrls.statements.cashflow_statement(),
)

period_cols = sorted(set(is_pc) & set(bs_pc) & set(cf_pc))
if not period_cols:
    raise SystemExit("No overlapping periods across IS / BS / CF — cannot compute ratios.")
fy_labels = [_fy_label(c) for c in period_cols]
step(f"Periods: {fy_labels}", t_start)

# ── Line item lookup ──────────────────────────────────────────────────────────
step("Looking up line items...", t_start)

if SHOW_ITEM_MATCHES:
    print("\n  -- IS matches --")
# std_concepts lists cover both US GAAP and IFRS names (edgartools partially normalises
# IFRS taxonomy to the same standard_concept strings, but a few diverge — listed below).
# IFRS divergences confirmed from TSM 20-F diagnostic:
#   Revenue (IFRS) vs Revenues (GAAP)
#   PretaxIncomeLoss (IFRS short form) vs IncomeLossFromContinuingOperations... (GAAP)
#   Equity (IFRS) vs StockholdersEquity (GAAP)
#   CurrentAssets / CurrentLiabilities (IFRS) vs AssetsCurrent / LiabilitiesCurrent (GAAP)
#   CostOfSales (IFRS) vs CostOfRevenue / CostOfGoodsSold (GAAP)
#   NoncurrentBorrowings (IFRS) vs LongTermDebt (GAAP)

revenue    = _get_item(is_df, period_cols,
                       ["Revenue", "Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax",
                        "SalesRevenueNet"],                             # Revenue = IFRS
                       ["net revenue", "total net revenue", "total revenue", "revenues", "net sales"],
                       _label="revenue")
gross_prof = _get_item(is_df, period_cols,
                       ["GrossProfit"],                                 # same in both
                       ["gross profit"],
                       _label="gross_prof")
op_income  = _get_item(is_df, period_cols,
                       None,   # sc matching unreliable: edgartools maps adjustment lines to OperatingIncomeLoss on some IFRS filings
                       ["income from operations", "operating income", "operating profit",
                        "profit from operations"],
                       _label="op_income")
net_income = _get_item(is_df, period_cols,
                       ["NetIncomeLoss", "ProfitLoss"],                 # ProfitLoss = IFRS
                       ["net income", "net earnings", "profit for the year",
                        "profit attributable"],
                       _label="net_income")
int_exp    = _get_item(is_df, period_cols,
                       ["InterestExpense", "InterestAndDebtExpense", "FinanceCosts"],
                       ["interest expense", "finance costs"],           # FinanceCosts = IFRS
                       _label="int_exp")
pretax     = _get_item(is_df, period_cols,
                       ["PretaxIncomeLoss",                             # IFRS short form
                        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
                        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments"],
                       ["income before income tax", "before income tax", "before provision",
                        "profit before tax"],
                       _label="pretax")
tax_exp    = _get_item(is_df, period_cols,
                       ["IncomeTaxExpense", "IncomeTaxExpenseBenefit"],
                       ["income tax expense", "tax expense", "provision for income tax"],
                       _label="tax_exp")
cogs       = _get_item(is_df, period_cols,
                       ["CostOfRevenue", "CostOfGoodsSold", "CostOfSales",
                        "CostOfGoodsSoldAndOccupancyCosts"],            # CostOfSales = IFRS
                       ["cost of revenue", "cost of sales", "cost of goods"],
                       _label="cogs")

# Accounting identity fallbacks — each uses the other two directly-found values.
# gross_prof first (uses revenue+cogs); revenue second (uses gross_prof+cogs).
if gross_prof is None and revenue is not None and cogs is not None:
    if SHOW_ITEM_MATCHES:
        print("  gross_prof           [fallback: revenue - cogs]")
    gross_prof = {
        c: (revenue[c] - cogs[c])
        if (revenue.get(c) is not None and pd.notna(revenue.get(c))
            and cogs.get(c) is not None and pd.notna(cogs.get(c)))
        else None
        for c in period_cols
    }

if revenue is None and gross_prof is not None and cogs is not None:
    if SHOW_ITEM_MATCHES:
        print("  revenue              [fallback: gross_prof + cogs]")
    revenue = {
        c: (gross_prof[c] + cogs[c])
        if (gross_prof.get(c) is not None and pd.notna(gross_prof.get(c))
            and cogs.get(c) is not None and pd.notna(cogs.get(c)))
        else None
        for c in period_cols
    }

if SHOW_ITEM_MATCHES:
    print("\n  -- BS matches --")
curr_assets = _get_item(bs_df, period_cols,
                        ["AssetsCurrent", "CurrentAssets"],            # CurrentAssets = IFRS
                        ["total current assets", "current assets"],
                        _label="curr_assets")
curr_liab   = _get_item(bs_df, period_cols,
                        ["LiabilitiesCurrent", "CurrentLiabilities"],  # CurrentLiabilities = IFRS
                        ["total current liabilities", "current liabilities"],
                        _label="curr_liab")
equity      = _get_item(bs_df, period_cols,
                        ["Equity",                                      # IFRS
                         "StockholdersEquity",
                         "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],
                        ["total equity", "total stockholders", "total shareholders"],
                        _label="equity")
cash        = _get_item(bs_df, period_cols,
                        ["CashAndCashEquivalentsAtCarryingValue",
                         "CashCashEquivalentsAndShortTermInvestments",
                         "CashAndCashEquivalents"],
                        ["cash and cash equivalents", "cash, cash equivalents"],
                        _label="cash")
st_debt     = _get_item(bs_df, period_cols,
                        ["ShortTermBorrowings", "CommercialPaper", "DebtCurrent",
                         "CurrentBorrowings"],                          # CurrentBorrowings = IFRS
                        ["short-term debt", "commercial paper", "current portion of long-term",
                         "current borrowings"],
                        _label="st_debt")
lt_debt     = _get_item(bs_df, period_cols,
                        ["LongTermDebt", "LongTermDebtNoncurrent",
                         "NoncurrentBorrowings", "NoncurrentPortionOfLongtermBorrowings"],
                        ["long-term debt", "term debt", "noncurrent borrowings"],
                        _label="lt_debt")
ar          = _get_item(bs_df, period_cols,
                        ["AccountsReceivableNetCurrent", "TradeAndOtherCurrentReceivables",
                         "TradeReceivables"],
                        ["accounts receivable", "trade receivables", "trade and other receivables",
                         "receivables, net", "net receivables"],
                        _label="ar")
inventory   = _get_item(bs_df, period_cols,
                        ["InventoryNet", "Inventories"],
                        ["inventories", "inventory"],
                        _label="inventory")
ap          = _get_item(bs_df, period_cols,
                        ["AccountsPayableCurrent", "TradeAndOtherCurrentPayables"],
                        ["accounts payable", "trade payables", "trade and other payables"],
                        _label="ap")

if SHOW_ITEM_MATCHES:
    print("\n  -- CF matches --")
cfo   = _get_item(cf_df, period_cols,
                  ["NetCashProvidedByUsedInOperatingActivities",
                   "CashFlowsFromUsedInOperatingActivities",            # IFRS
                   "NetCashFromOperatingActivities"],
                  ["operating activities", "cash from operations", "cash provided by operating",
                   "cash generated by operating"],
                  _label="cfo")
capex = _get_item(cf_df, period_cols,
                  ["PaymentsToAcquirePropertyPlantAndEquipment",
                   "PurchaseOfPropertyPlantAndEquipment",               # IFRS singular
                   "PurchasesOfPropertyPlantAndEquipment",             # IFRS plural
                   "CapitalExpenses"],                                  # edgartools mapped
                  ["capital expenditure",                               # singular (UK/IFRS)
                   "capital expenditures",                              # plural (US GAAP)
                   "expenditures for property",                         # MU-style
                   "purchase of property", "purchases of property",
                   "acquisition of property", "acquisitions of property"],
                  exclude_patterns=["included in accrued", "non-cash"],
                  _label="capex")
da    = _get_item(cf_df, period_cols,
                  ["DepreciationDepletionAndAmortization", "DepreciationAndAmortization",
                   "DepreciationExpense",                               # edgartools mapped (MU)
                   "AdjustmentsForDepreciationAndAmortisationExpense",  # IFRS British
                   "DepreciationAmortisationAndImpairmentLossRecognisedInProfitOrLoss"],
                  ["depreciation and amortization",                     # US spelling
                   "depreciation and amortisation",                     # British/IFRS spelling
                   "depreciation, amortization", "depreciation, amortisation",
                   "depreciation and depletion",
                   "depreciation expense and amortization",             # MU-style
                   "depreciation expense",                              # MU/generic fallback
                   "depreciation of property and equipment",            # GOOG-style split line
                   "depreciation of property"],
                  _label="da")

# ── Market cap ────────────────────────────────────────────────────────────────
step("Fetching market cap from yfinance...", t_start)
mkt_cap_usd = None
try:
    mkt_cap_usd = yf.Ticker(TICKER).fast_info.market_cap
    step(f"Market cap: ${mkt_cap_usd / 1e9:.1f}B", t_start)
except Exception as e:
    step(f"Market cap unavailable: {e}", t_start)

mkt_cap_bn = mkt_cap_usd / 1e9 if mkt_cap_usd else None

# ── Ratio computation ─────────────────────────────────────────────────────────
step("Computing ratios...", t_start)

gross_margin = _ratio(gross_prof, revenue, period_cols)
op_margin    = _ratio(op_income,  revenue, period_cols)
net_margin   = _ratio(net_income, revenue, period_cols)

# ROIC = NOPAT / Invested Capital
# NOPAT = EBIT * (1 - effective tax rate); Invested Capital = Equity + Net Debt
tax_rate = _ratio(tax_exp, pretax, period_cols)
# Fallback: derive effective tax rate from 1 - (net_income / pretax) when tax_exp not found
if tax_exp is None and net_income and pretax:
    tax_rate = {
        c: (1 - net_income[c] / pretax[c])
        if (pd.notna(net_income.get(c)) and pd.notna(pretax.get(c))
            and pretax.get(c) not in (0, None))
        else None
        for c in period_cols
    }
nopat: dict = {}
for c in period_cols:
    oi = (op_income or {}).get(c)
    tr = tax_rate.get(c)
    nopat[c] = oi * (1 - min(max(tr, 0), 0.5)) if pd.notna(oi) and pd.notna(tr) else None

total_debt       = _add(st_debt, lt_debt, period_cols, allow_partial=True)
net_debt         = {c: (total_debt.get(c) or 0) - (cash.get(c) or 0)
                   if total_debt.get(c) is not None or cash.get(c) is not None else None
                   for c in period_cols}
invested_capital = _add(equity, net_debt, period_cols, allow_partial=True)
roic             = _ratio(nopat, invested_capital, period_cols)

avg_equity = _avg_adjacent(equity, period_cols) if equity else None
roe        = _ratio(net_income, avg_equity, period_cols)

# EBITDA = Operating Income + D&A (add-back)
ebitda: dict = {}
for c in period_cols:
    oi = (op_income or {}).get(c)
    d  = (da or {}).get(c)
    if pd.notna(oi):
        ebitda[c] = oi + abs(d) if pd.notna(d) else oi
    else:
        ebitda[c] = None

nd_ebitda    = _ratio(net_debt, ebitda, period_cols)
abs_int_exp  = {c: abs(v) if pd.notna(v) else None for c, v in (int_exp or {}).items()} if int_exp else None
int_coverage = _ratio(op_income, abs_int_exp, period_cols)

current_ratio = _ratio(curr_assets, curr_liab, period_cols)
dso           = _ratio(ar,        revenue, period_cols, scale=365)
dio           = _ratio(inventory, cogs,    period_cols, scale=365)
dpo           = _ratio(ap,        cogs,    period_cols, scale=365)

ccc: dict = {}
for c in period_cols:
    d, i, p = dso.get(c), dio.get(c), dpo.get(c)
    if pd.notna(d) and pd.notna(p):
        ccc[c] = d + (i if pd.notna(i) else 0) - p
    else:
        ccc[c] = None

fcf: dict = {}
for c in period_cols:
    c_val  = (cfo   or {}).get(c)
    cx_val = (capex or {}).get(c)
    if pd.notna(c_val) and pd.notna(cx_val):
        fcf[c] = c_val - abs(cx_val)   # abs() handles sign ambiguity
    elif pd.notna(c_val):
        fcf[c] = c_val
    else:
        fcf[c] = None

fcf_margin = _ratio(fcf, revenue, period_cols)
fcf_yield  = {
    c: (fcf.get(c) * 1_000_000) / mkt_cap_usd
    if mkt_cap_usd and fcf.get(c) is not None and pd.notna(fcf.get(c)) else None
    for c in period_cols
}

ratios = {
    "gross_margin":  gross_margin,
    "op_margin":     op_margin,
    "net_margin":    net_margin,
    "roic":          roic,
    "roe":           roe,
    "nd_ebitda":     nd_ebitda,
    "int_coverage":  int_coverage,
    "current_ratio": current_ratio,
    "dso":           dso,
    "dio":           dio,
    "dpo":           dpo,
    "ccc":           ccc,
    "fcf":           fcf,
    "fcf_margin":    fcf_margin,
    "fcf_yield":     fcf_yield,
}

# ── Console preview ───────────────────────────────────────────────────────────
print(f"\n--- Financial Ratios ({form_type}, {TICKER}) ---", flush=True)
preview_rows = []
for section in RATIO_SECTIONS:
    preview_rows.append({"Ratio": f"  {section['title'].upper()}"})
    for rdef in section["ratios"]:
        row = {"Ratio": f"  {rdef['label']}"}
        for c, fy in zip(period_cols, fy_labels):
            v = ratios.get(rdef["key"], {}).get(c)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                row[fy] = "N/A"
            elif rdef["fmt"] == FMT_PCT:
                row[fy] = f"{v:.1%}"
            elif rdef["fmt"] == FMT_X:
                row[fy] = f"{v:.1f}x"
            else:
                row[fy] = f"{v:,.0f}"
        preview_rows.append(row)

preview = pd.DataFrame(preview_rows).fillna("")
print(preview.to_string(index=False), flush=True)

# ── Excel export ──────────────────────────────────────────────────────────────
out_dir  = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{TICKER}_ratios.xlsx"

step(f"Writing Excel workbook to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)
write_ratio_sheet(wb, ratios, period_cols, fy_labels, company.name, form_type, mkt_cap_bn)
wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel.", t_start)
