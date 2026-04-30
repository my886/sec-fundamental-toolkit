r"""
96_comps.py — Comparable companies table

Build a side-by-side equity research comps sheet for a peer group.
One column per company, latest fiscal year metrics as rows.

Metric groups:
  Trading:        Market Cap ($B), Enterprise Value ($B)
  Income:         Revenue ($M), Rev Growth YoY, Gross Margin, EBIT Margin, Net Margin, EBITDA ($M)
  Returns:        ROIC, ROE
  Capital Struct: Net Debt ($M), Net Debt / EBITDA
  Valuation:      EV / Revenue, EV / EBITDA, FCF Yield *
  FCF:            FCF ($M), FCF Margin

Note: 20-F filers (non-US issuers) may report in their local currency.
  Absolute values are then in that currency; EV multiples are marked N/A.

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\96_comps.py
"""
import os
import re
import time
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
os.environ["EDGAR_IDENTITY"] = os.getenv("SEC_USER_AGENT", "")

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from edgar import Company
from edgar.xbrl import XBRLS

# ── Configuration ─────────────────────────────────────────────────────────────
PEER_GROUP = "semiconductors"
TICKERS    = ["NVDA", "AVGO", "TSM", "AMD", "QCOM", "TXN", "MU", "INTC", "AMAT", "LRCX"]
YEARS      = 5   # number of filings to stitch; only latest FY is used for comps

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"
CLR_HDR_BG     = "4472C4"
CLR_SECTION_BG = "D9E1F2"
CLR_ALT_BG     = "F9F9F9"
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"
CLR_NA         = "AAAAAA"
THIN = Side(style="thin",   color="CCCCCC")
MED  = Side(style="medium", color="AAAAAA")

# ── Comps table layout ────────────────────────────────────────────────────────
COMP_SECTIONS = [
    {
        "title": "Trading Data",
        "rows": [
            {"label": "Market Cap ($B)",      "key": "mkt_cap_bn",  "fmt": "#,##0.0"},
            {"label": "Enterprise Value ($B)", "key": "ev_bn",       "fmt": "#,##0.0"},
        ],
    },
    {
        "title": "Income Statement",
        "rows": [
            {"label": "Revenue ($M)",    "key": "revenue",      "fmt": "#,##0"},
            {"label": "Revenue Growth",  "key": "rev_growth",   "fmt": "0.0%"},
            {"label": "Gross Margin",    "key": "gross_margin", "fmt": "0.0%"},
            {"label": "EBIT Margin",     "key": "op_margin",    "fmt": "0.0%"},
            {"label": "Net Margin",      "key": "net_margin",   "fmt": "0.0%"},
            {"label": "EBITDA ($M)",     "key": "ebitda",       "fmt": "#,##0"},
        ],
    },
    {
        "title": "Returns",
        "rows": [
            {"label": "ROIC", "key": "roic", "fmt": "0.0%"},
            {"label": "ROE",  "key": "roe",  "fmt": "0.0%"},
        ],
    },
    {
        "title": "Capital Structure",
        "rows": [
            {"label": "Net Debt ($M)",     "key": "net_debt",  "fmt": "#,##0"},
            {"label": "Net Debt / EBITDA", "key": "nd_ebitda", "fmt": '0.0"x"'},
        ],
    },
    {
        "title": "Valuation",
        "rows": [
            {"label": "EV / Revenue",  "key": "ev_rev",    "fmt": '0.0"x"'},
            {"label": "EV / EBITDA",   "key": "ev_ebitda", "fmt": '0.0"x"'},
            {"label": "FCF Yield *",   "key": "fcf_yield", "fmt": "0.0%"},
        ],
    },
    {
        "title": "Free Cash Flow",
        "rows": [
            {"label": "FCF ($M)",   "key": "fcf",        "fmt": "#,##0"},
            {"label": "FCF Margin", "key": "fcf_margin", "fmt": "0.0%"},
        ],
    },
]


# ── Utilities ─────────────────────────────────────────────────────────────────

def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


def _period_cols(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if re.match(r"\d{4}-\d{2}-\d{2}$", c)]


# ── Data loading (same pattern as 97_ratios.py / 98_multiperiod_financials.py) ─

def _prepare_structured_df(template_stmt, stitched_stmt) -> tuple[pd.DataFrame, list[str]]:
    """Merge single-filing structure with XBRLS multi-period values.
    Falls back to single-filing data if stitching raises an exception."""
    tmpl = template_stmt.to_dataframe(view="standard")
    try:
        stitched = stitched_stmt.to_dataframe()
    except Exception:
        stitched = tmpl.copy()
        rename = {
            col: re.match(r"(\d{4}-\d{2}-\d{2})", col).group(1)
            for col in stitched.columns
            if re.match(r"\d{4}-\d{2}-\d{2}", col) and not re.match(r"\d{4}-\d{2}-\d{2}$", col)
        }
        stitched = stitched.rename(columns=rename)
    period_cols = sorted(_period_cols(stitched))

    for c in period_cols:
        stitched[c] = pd.to_numeric(stitched[c], errors="coerce")

    stitch_lookup: dict = {}
    for _, srow in stitched.iterrows():
        concept = srow["concept"]
        if concept not in stitch_lookup:
            stitch_lookup[concept] = {c: srow[c] for c in period_cols}

    rows = []
    for _, trow in tmpl.iterrows():
        if trow.get("dimension", False):
            continue
        concept = trow["concept"]
        sc      = trow.get("standard_concept")
        is_abs  = bool(trow.get("abstract", False))
        row_data = {
            "label":            trow["label"],
            "level":            int(trow.get("level", 1)),
            "abstract":         is_abs,
            "standard_concept": sc if pd.notna(sc) else "",
        }
        if is_abs:
            for c in period_cols:
                row_data[c] = None
        else:
            hist = stitch_lookup.get(concept, {})
            for c in period_cols:
                row_data[c] = hist.get(c)
        rows.append(row_data)

    df = pd.DataFrame(rows)
    data_mask = ~df["abstract"]
    df = df[df["abstract"] | df.loc[data_mask, period_cols].notna().any(axis=1)].reset_index(drop=True)

    for idx, row in df.iterrows():
        if row["abstract"]:
            continue
        vals = [row[c] for c in period_cols if pd.notna(row[c])]
        if vals and max(abs(v) for v in vals) >= 1_000:
            for c in period_cols:
                if pd.notna(df.at[idx, c]):
                    df.at[idx, c] = df.at[idx, c] / 1_000_000

    return df, period_cols


def _get_item(
    df: pd.DataFrame,
    period_cols: list[str],
    std_concepts: list[str] | None = None,
    label_patterns: list[str] | None = None,
    exclude_patterns: list[str] | None = None,
) -> dict | None:
    """Return {period_date: value} for best-matching non-abstract row, or None."""
    data = df[~df["abstract"]]
    if exclude_patterns:
        excl = pd.Series(False, index=data.index)
        for pat in exclude_patterns:
            excl |= data["label"].str.lower().str.contains(pat.lower(), na=False)
        data = data[~excl]

    def _best(matches):
        m = matches.copy()
        m["_llen"]   = m["label"].str.len()
        # "Other ..." rows are supplemental disclosures; prefer non-"Other" at same level
        m["_is_other"] = m["label"].str.strip().str.lower().str.startswith("other").astype(int)
        return m.sort_values(["level", "_is_other", "_llen"]).iloc[0]

    def _result(row):
        return {c: pd.to_numeric(row.get(c), errors="coerce") for c in period_cols}

    if std_concepts:
        for sc in std_concepts:
            mask = data["standard_concept"].str.lower() == sc.lower()
            if mask.any():
                return _result(_best(data[mask]))

    if label_patterns:
        for pat in label_patterns:
            mask = data["label"].str.lower().str.contains(pat.lower(), na=False)
            if mask.any():
                m = data[mask].copy()
                m["_llen"]     = m["label"].str.len()
                m["_is_other"] = m["label"].str.strip().str.lower().str.startswith("other").astype(int)
                return _result(m.sort_values(["level", "_is_other", "_llen"]).iloc[0])

    return None


# ── Per-ticker fetch ───────────────────────────────────────────────────────────

def fetch_ticker_data(ticker: str, years: int) -> dict | None:
    """Fetch and compute latest-FY metrics for one ticker. Returns None on failure."""
    try:
        company = Company(ticker)
        form_type = None
        filings   = None
        for form in ("10-K", "20-F"):
            candidate = company.get_filings(form=form, amendments=False).head(years)
            if len(candidate) > 0:
                form_type = form
                filings   = candidate
                break
        if not filings:
            print(f"  [{ticker}] No 10-K or 20-F filings found.", flush=True)
            return None

        template_fin = company.get_financials()
        if template_fin is None:
            print(f"  [{ticker}] Could not parse financials.", flush=True)
            return None

        xbrls = XBRLS.from_filings(filings)

        is_df, is_pc = _prepare_structured_df(template_fin.income_statement(),   xbrls.statements.income_statement())
        bs_df, bs_pc = _prepare_structured_df(template_fin.balance_sheet(),      xbrls.statements.balance_sheet())
        cf_df, cf_pc = _prepare_structured_df(template_fin.cashflow_statement(), xbrls.statements.cashflow_statement())

        period_cols = sorted(set(is_pc) & set(bs_pc) & set(cf_pc))
        if not period_cols:
            print(f"  [{ticker}] No overlapping periods across IS / BS / CF.", flush=True)
            return None

        latest = period_cols[-1]
        prev   = period_cols[-2] if len(period_cols) >= 2 else None

        # IS
        revenue    = _get_item(is_df, period_cols,
                               ["Revenue", "Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax", "SalesRevenueNet"],
                               ["net revenue", "total net revenue", "total revenue", "revenues", "net sales"])
        gross_prof = _get_item(is_df, period_cols, ["GrossProfit"], ["gross profit"])
        op_income  = _get_item(is_df, period_cols, None,
                               ["income from operations", "operating income", "operating profit", "profit from operations"])
        net_income = _get_item(is_df, period_cols, ["NetIncomeLoss", "ProfitLoss"],
                               ["net income", "net earnings", "profit for the year", "profit attributable"])
        pretax     = _get_item(is_df, period_cols,
                               ["PretaxIncomeLoss",
                                "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest"],
                               ["income before income tax", "before income tax", "profit before tax"])
        tax_exp    = _get_item(is_df, period_cols, ["IncomeTaxExpense", "IncomeTaxExpenseBenefit"],
                               ["income tax expense", "tax expense", "provision for income tax"])
        cogs       = _get_item(is_df, period_cols,
                               ["CostOfRevenue", "CostOfGoodsSold", "CostOfSales",
                                "CostOfGoodsAndServicesSold", "CostOfGoodsSoldAndOccupancyCosts"],
                               ["cost of revenue", "cost of sales", "cost of goods"])
        int_exp    = _get_item(is_df, period_cols,
                               ["InterestExpense", "InterestAndDebtExpense", "FinanceCosts"],
                               ["interest expense", "finance costs"])

        # Identity fallbacks
        if gross_prof is None and revenue is not None and cogs is not None:
            gross_prof = {c: (revenue[c] - cogs[c])
                          if pd.notna(revenue.get(c)) and pd.notna(cogs.get(c)) else None
                          for c in period_cols}
        if revenue is None and gross_prof is not None and cogs is not None:
            revenue = {c: (gross_prof[c] + cogs[c])
                       if pd.notna(gross_prof.get(c)) and pd.notna(cogs.get(c)) else None
                       for c in period_cols}
        # Some IS structures omit an explicit operating income subtotal (e.g. KLAC).
        # Use pretax as a last resort — close enough when net interest is small.
        if op_income is None and pretax is not None:
            op_income = pretax

        # BS
        equity  = _get_item(bs_df, period_cols,
                            ["Equity", "StockholdersEquity",
                             "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],
                            ["total equity", "total stockholders", "total shareholders"])
        cash    = _get_item(bs_df, period_cols,
                            ["CashAndCashEquivalentsAtCarryingValue",
                             "CashCashEquivalentsAndShortTermInvestments",
                             "CashAndCashEquivalents"],
                            ["cash and cash equivalents", "cash, cash equivalents"])
        st_debt = _get_item(bs_df, period_cols,
                            ["ShortTermBorrowings", "CommercialPaper", "DebtCurrent", "CurrentBorrowings"],
                            ["short-term debt", "commercial paper", "current portion of long-term",
                             "current borrowings"])
        lt_debt = _get_item(bs_df, period_cols,
                            ["LongTermDebt", "LongTermDebtNoncurrent",
                             "NoncurrentBorrowings", "NoncurrentPortionOfLongtermBorrowings"],
                            ["long-term debt", "term debt", "noncurrent borrowings"])

        # CF
        cfo   = _get_item(cf_df, period_cols,
                          ["NetCashProvidedByUsedInOperatingActivities",
                           "CashFlowsFromUsedInOperatingActivities",
                           "NetCashFromOperatingActivities"],
                          ["operating activities", "cash from operations",
                           "cash provided by operating", "cash generated by operating"])
        capex = _get_item(cf_df, period_cols,
                          ["PaymentsToAcquirePropertyPlantAndEquipment",
                           "PurchaseOfPropertyPlantAndEquipment",
                           "PurchasesOfPropertyPlantAndEquipment",
                           "CapitalExpenses"],
                          ["capital expenditure", "capital expenditures",
                           "expenditures for property",
                           "purchase of property", "purchases of property",
                           "acquisition of property", "acquisitions of property"],
                          exclude_patterns=["included in accrued", "non-cash"])
        da    = _get_item(cf_df, period_cols,
                          ["DepreciationDepletionAndAmortization", "DepreciationAndAmortization",
                           "DepreciationExpense",
                           "AdjustmentsForDepreciationAndAmortisationExpense",
                           "DepreciationAmortisationAndImpairmentLossRecognisedInProfitOrLoss"],
                          ["depreciation and amortization", "depreciation and amortisation",
                           "depreciation, amortization", "depreciation, amortisation",
                           "depreciation and depletion",
                           "depreciation expense and amortization", "depreciation expense",
                           "depreciation of property and equipment", "depreciation of property"])

        # Market cap
        mkt_cap_usd = None
        try:
            mkt_cap_usd = yf.Ticker(ticker).fast_info.market_cap
        except Exception:
            pass


        # ── Single-value helper ───────────────────────────────────────────────
        def v(d, period=latest):
            val = (d or {}).get(period)
            return val if (val is not None and pd.notna(val)) else None

        def pct(num, den):
            return num / den if (num is not None and den is not None and den != 0) else None

        # ── Compute metrics for latest FY ─────────────────────────────────────
        rev_v   = v(revenue)
        gp_v    = v(gross_prof)
        oi_v    = v(op_income)
        ni_v    = v(net_income)
        da_v    = v(da)
        eq_v    = v(equity)
        cash_v  = v(cash) or 0
        std_v   = v(st_debt) or 0
        ltd_v   = v(lt_debt) or 0
        cfo_v   = v(cfo)
        capex_v = v(capex)
        pt_v    = v(pretax)
        tax_v   = v(tax_exp)

        prev_rev_v = v(revenue, prev) if prev else None
        prev_eq_v  = v(equity,  prev) if prev else None

        total_debt_v = (std_v + ltd_v) if (st_debt is not None or lt_debt is not None) else None
        net_debt_v   = (total_debt_v - cash_v) if total_debt_v is not None else None
        ebitda_v     = (oi_v + abs(da_v)) if oi_v is not None and da_v is not None else oi_v

        fcf_v = (cfo_v - abs(capex_v)) if cfo_v is not None and capex_v is not None else cfo_v

        # Tax rate — prefer explicit, fall back to 1 - (NI / Pretax)
        if tax_v is not None and pt_v and pt_v != 0:
            tax_rate = tax_v / pt_v
        elif ni_v is not None and pt_v and pt_v != 0:
            tax_rate = 1 - ni_v / pt_v
        else:
            tax_rate = None

        nopat_v = oi_v * (1 - min(max(tax_rate, 0), 0.5)) if oi_v is not None and tax_rate is not None else None
        avg_eq_v = (eq_v + prev_eq_v) / 2 if eq_v is not None and prev_eq_v is not None else eq_v
        invested_capital_v = (eq_v + (net_debt_v or 0)) if eq_v is not None else None

        # EV: only valid for USD filers (10-K); 20-F may report in local currency
        mkt_cap_m = mkt_cap_usd / 1e6 if mkt_cap_usd else None
        is_usd    = (form_type == "10-K")
        ev_m      = (mkt_cap_m + (net_debt_v or 0)) if mkt_cap_m is not None and is_usd else None

        return {
            "ticker":      ticker,
            "name":        company.name,
            "fy_end":      latest,
            "form_type":   form_type,
            "is_usd":      is_usd,
            "mkt_cap_bn":  mkt_cap_usd / 1e9 if mkt_cap_usd else None,
            "ev_bn":       ev_m / 1000 if ev_m is not None else None,
            "revenue":     rev_v,
            "rev_growth":  pct(rev_v - prev_rev_v, prev_rev_v) if rev_v is not None and prev_rev_v else None,
            "gross_margin": pct(gp_v, rev_v),
            "op_margin":   pct(oi_v, rev_v),
            "net_margin":  pct(ni_v, rev_v),
            "ebitda":      ebitda_v,
            "roic":        pct(nopat_v, invested_capital_v),
            "roe":         pct(ni_v, avg_eq_v),
            "net_debt":    net_debt_v,
            "nd_ebitda":   pct(net_debt_v, ebitda_v) if ebitda_v and ebitda_v > 0 else None,
            "ev_rev":      pct(ev_m, rev_v),
            "ev_ebitda":   pct(ev_m, ebitda_v) if ebitda_v and ebitda_v > 0 else None,
            "fcf_yield":   pct(fcf_v, mkt_cap_m) if is_usd else None,
            "fcf":         fcf_v,
            "fcf_margin":  pct(fcf_v, rev_v),
        }

    except Exception as e:
        print(f"  [{ticker}] Failed: {e}", flush=True)
        return None


# ── Excel output ──────────────────────────────────────────────────────────────

def write_comps_sheet(wb: Workbook, results: list[dict]) -> None:
    n = len(results)
    n_cols = 1 + n   # label column + one per company

    ws = wb.create_sheet(title="Comps")

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.append([f"Peer Group: {PEER_GROUP.upper()}  —  Comparable Companies"] + [""] * n)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1)
    tc.font      = Font(bold=True, color=CLR_WHITE, size=12)
    tc.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ── Company header row ────────────────────────────────────────────────────
    hdr = ["Metric"]
    for r in results:
        note = "" if r["is_usd"] else f"\n({r['form_type']})"
        hdr.append(f"{r['ticker']}\nFY{r['fy_end'][:4]}{note}")
    ws.append(hdr)
    for col in range(1, n_cols + 1):
        cell = ws.cell(2, col)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HDR_BG)
        cell.alignment = Alignment(
            horizontal="right" if col > 1 else "left",
            vertical="center", wrap_text=True,
        )
        cell.border = Border(bottom=MED)
    ws.row_dimensions[2].height = 40

    # ── Data rows ─────────────────────────────────────────────────────────────
    odd = True
    for section in COMP_SECTIONS:
        # Section header
        ws.append([section["title"]] + [""] * n)
        hr = ws.max_row
        for col in range(1, n_cols + 1):
            cell = ws.cell(hr, col)
            cell.fill      = PatternFill("solid", fgColor=CLR_SECTION_BG)
            cell.border    = Border(bottom=THIN)
            cell.font      = Font(bold=True, size=10, color=CLR_DARK)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[hr].height = 16

        for rdef in section["rows"]:
            row_vals = [rdef["label"]] + [r.get(rdef["key"]) for r in results]
            ws.append(row_vals)
            dr  = ws.max_row
            bg  = PatternFill("solid", fgColor=CLR_ALT_BG) if odd else None
            odd = not odd

            for col in range(1, n_cols + 1):
                cell = ws.cell(dr, col)
                if bg:
                    cell.fill = bg
                cell.border = Border(bottom=THIN)

                if col == 1:
                    cell.font      = Font(size=10, color=CLR_DARK)
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
                else:
                    val = row_vals[col - 1]
                    if val is not None and pd.notna(val):
                        cell.number_format = rdef["fmt"]
                        cell.font          = Font(size=10, color=CLR_DARK)
                    else:
                        cell.value = "N/A"
                        cell.font  = Font(size=10, color=CLR_NA, italic=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            ws.row_dimensions[dr].height = 15

    # ── Footnotes ─────────────────────────────────────────────────────────────
    ws.append([""])
    footnotes = [
        "* FCF Yield = FCF / Market Cap (current, from yfinance). Uses latest-available mkt cap, not period-end mkt cap.",
        "  EV = Market Cap + Net Debt. EV metrics shown for 10-K (USD) filers only.",
    ]
    if any(not r["is_usd"] for r in results):
        footnotes.append(
            "  20-F filers report in their local currency; absolute values (Revenue, EBITDA, "
            "FCF, Net Debt) are in that currency."
        )
    for fn in footnotes:
        ws.append([fn] + [""] * n)
        fn_row = ws.max_row
        ws.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=n_cols)
        fn_cell = ws.cell(fn_row, 1)
        fn_cell.font      = Font(italic=True, size=8, color=CLR_NA)
        fn_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Column widths + freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    col_w = max(10, min(16, int(90 / n)))  # narrower for larger peer groups
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = col_w
    ws.freeze_panes = "B3"


# ── Main ──────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

results = []
for ticker in TICKERS:
    step(f"Processing {ticker}...", t_start)
    data = fetch_ticker_data(ticker, YEARS)
    if data:
        results.append(data)
        step(f"  {ticker} done — {data['name']} FY{data['fy_end'][:4]} ({data['form_type']})", t_start)
    else:
        step(f"  {ticker} skipped.", t_start)

if not results:
    raise SystemExit("No data retrieved for any ticker.")

# Sort by market cap descending (standard for equity research comps)
results.sort(key=lambda r: r.get("mkt_cap_bn") or 0, reverse=True)

# Console preview
print(f"\n--- {PEER_GROUP.upper()} Comps (latest FY) ---", flush=True)
hdr_line = f"{'Metric':<22}" + "".join(f"{r['ticker']:>12}" for r in results)
print(hdr_line, flush=True)
print("-" * len(hdr_line), flush=True)
for section in COMP_SECTIONS:
    print(f"  {section['title'].upper()}", flush=True)
    for rdef in section["rows"]:
        vals = []
        for r in results:
            val = r.get(rdef["key"])
            if val is None or (isinstance(val, float) and pd.isna(val)):
                vals.append("N/A")
            elif rdef["fmt"].endswith('%"'):
                vals.append(f"{val:.1%}")
            elif rdef["fmt"] == "0.0%":
                vals.append(f"{val:.1%}")
            elif '"x"' in rdef["fmt"]:
                vals.append(f"{val:.1f}x")
            else:
                vals.append(f"{val:,.0f}")
        print(f"  {rdef['label']:<20}" + "".join(f"{v:>12}" for v in vals), flush=True)

# Excel export
out_dir  = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{PEER_GROUP}_comps.xlsx"

step(f"\nWriting Excel to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)
write_comps_sheet(wb, results)
wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel.", t_start)
