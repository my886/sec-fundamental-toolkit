r"""
Pull last N quarters of 10-Q data for a single ticker and export per-statement
Excel sheets with:
  - Quarterly values (oldest left, newest right)
  - LTM: latest annual + post-annual quarters - same quarters prior year
  - Balance sheet: Latest Q value (point-in-time, no LTM sum)
  - QoQ% and YoY% growth per line item

LTM formula (per BACKLOG spec):
  LTM = last 10-K annual + stub quarters since 10-K end - same quarters prior year
  Falls back to 4-quarter sum when annual data is unavailable.
  Balance sheet: always Latest Q (point-in-time, not summable).

Note: Q4 is reported in the 10-K, not in 10-Q filings. This script fetches
each 10-Q individually (not via XBRLS) to get the standalone (Qn) period
per quarter. XBRLS is not used here because its stitched output uses plain
YYYY-MM-DD columns with no quarter suffix, making Q vs YTD indistinguishable.

Foreign issuers (ASML, TSM) file 6-K for interim reports, but most 6-Ks are
press releases without XBRL. Quarterly analysis is not available for these
companies via this script.

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\95_quarterly.py
"""
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
os.environ["EDGAR_IDENTITY"] = os.getenv("SEC_USER_AGENT", "")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from edgar import Company
from edgar.financials import Financials

# ── Configuration ─────────────────────────────────────────────────────────────
TICKER     = "KLAC" # KO, KLAC
N_QUARTERS = 8    # quarterly periods to display
N_FILINGS  = N_QUARTERS + 4  # total filings fetched; extra 4 needed for prior-year LTM lookup

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"
CLR_HDR_BG     = "4472C4"
CLR_SECTION_BG = "D9E1F2"
CLR_TOTAL_BG   = "EBF0FA"
CLR_LTM_BG     = "E2EFDA"   # light green — LTM column
CLR_ALT_BG     = "F9F9F9"
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"
THIN           = Side(style="thin",   color="CCCCCC")
MED            = Side(style="medium", color="AAAAAA")

TOTAL_KEYWORDS = (
    "total", "gross margin", "operating income", "net income",
    "earnings before", "ebitda", "free cash flow", "gross profit",
)

COL_LTM = "LTM"
COL_LQ  = "Latest Q"
COL_QOQ = "QoQ %"
COL_YOY = "YoY %"
PCT_FMT = "+0.0%;-0.0%;--"


def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


# ── Period column helpers ─────────────────────────────────────────────────────

def _q_col(df: pd.DataFrame) -> str | None:
    """Most recent standalone-quarter column — format '2024-06-30 (Q2)'."""
    cols = sorted([c for c in df.columns if re.search(r"\(Q[1-4]\)$", c)], reverse=True)
    return cols[0] if cols else None


def _ytd_col(df: pd.DataFrame) -> str | None:
    """Most recent YTD cumulative column — format '2024-12-31 (YTD)'.
    Cash flow Q2/Q3 10-Qs use YTD labels instead of (Q2)/(Q3)."""
    cols = sorted([c for c in df.columns if re.search(r"\(YTD\)$", c)], reverse=True)
    return cols[0] if cols else None


def _instant_col(df: pd.DataFrame) -> str | None:
    """Most recent plain-date column — balance sheet instant period."""
    cols = sorted([c for c in df.columns if re.match(r"\d{4}-\d{2}-\d{2}$", c)], reverse=True)
    return cols[0] if cols else None


def _fy_col(df: pd.DataFrame) -> str | None:
    """Most recent (FY) column from an annual statement."""
    cols = sorted([c for c in df.columns if re.search(r"\(FY\)$", c)], reverse=True)
    return cols[0] if cols else None


def _end_date(col: str) -> str | None:
    """Extract 'YYYY-MM-DD' from a column name."""
    m = re.match(r"(\d{4}-\d{2}-\d{2})", col)
    return m.group(1) if m else None


def _q_label(date_str: str) -> str:
    """'2024-06-30' -> 'Jun 2024'"""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%b %Y")
    except ValueError:
        return date_str


def _extract_values(df: pd.DataFrame, col: str) -> dict:
    """{concept: raw_numeric_value} for one period column, skipping abstract/dimension rows."""
    result = {}
    for _, row in df.iterrows():
        if row.get("abstract", False) or row.get("dimension", False):
            continue
        concept = row.get("concept")
        if not concept:
            continue
        val = pd.to_numeric(row[col], errors="coerce")
        if pd.notna(val):
            result[concept] = val
    return result


def _find_prior_date(date_str: str, candidates: list[str], tol_days: int = 45) -> str | None:
    """Return the candidate date closest to date_str - 365 days, within tol_days."""
    target = datetime.strptime(date_str, "%Y-%m-%d") - timedelta(days=365)
    best, best_diff = None, float("inf")
    for d in candidates:
        diff = abs((datetime.strptime(d, "%Y-%m-%d") - target).days)
        if diff <= tol_days and diff < best_diff:
            best, best_diff = d, diff
    return best


def _is_total_row(label: str) -> bool:
    low = label.strip().lower()
    return any(kw in low for kw in TOTAL_KEYWORDS)


def _num_fmt(row: pd.Series, cols: list[str]) -> str:
    vals = [row[c] for c in cols if c in row.index and pd.notna(row[c])]
    if not vals:
        return "#,##0"
    return "#,##0.00" if max(abs(v) for v in vals) < 1_000 else "#,##0"


# ── Core data preparation ─────────────────────────────────────────────────────

def _prepare_quarterly_df(
    template_stmt,
    q_lookup: dict,        # {concept: {date_str: raw_value}}
    all_q_dates: list,     # all available quarterly dates (for prior-year lookup)
    ann_lookup: dict,      # {concept: raw_value} from latest annual filing
    annual_end_date: str | None,
    is_balance_sheet: bool,
) -> tuple[pd.DataFrame, list[str], str]:
    """
    Build merged DataFrame:
      structure from template_stmt + values from q_lookup + LTM/QoQ/YoY.
    Returns (df, q_display_cols, ltm_col_name).
    """
    tmpl = template_stmt.to_dataframe(view="standard")
    q_display_cols = sorted(all_q_dates)[-N_QUARTERS:]

    # Stub and prior dates for LTM (income/CF only)
    if not is_balance_sheet and annual_end_date:
        ann_dt = datetime.strptime(annual_end_date, "%Y-%m-%d")
        stub_dates  = sorted([d for d in all_q_dates
                               if datetime.strptime(d, "%Y-%m-%d") > ann_dt])
        prior_dates = [_find_prior_date(d, all_q_dates) for d in stub_dates]
    else:
        stub_dates, prior_dates = [], []

    # Walk template for filing-order structure
    rows = []
    for _, trow in tmpl.iterrows():
        if trow.get("dimension", False):
            continue
        concept     = trow["concept"]
        sc          = trow.get("standard_concept")
        is_abstract = bool(trow.get("abstract", False))
        level       = int(trow.get("level", 1))

        row_data = {
            "concept":          concept,
            "label":            trow["label"],
            "level":            level,
            "abstract":         is_abstract,
            "standard_concept": sc if pd.notna(sc) else "",
        }
        if is_abstract:
            for c in q_display_cols:
                row_data[c] = None
        else:
            hist = q_lookup.get(concept, {})
            for c in q_display_cols:
                raw = hist.get(c)
                row_data[c] = pd.to_numeric(raw, errors="coerce") if raw is not None else None

        rows.append(row_data)

    df = pd.DataFrame(rows)

    # Drop all-NaN non-abstract rows
    data_mask = ~df["abstract"]
    if q_display_cols:
        df = df[df["abstract"] | df.loc[data_mask, q_display_cols].notna().any(axis=1)]
    df = df.reset_index(drop=True)

    # Scale to USD millions; track which concepts were scaled for consistent LTM scaling
    scaled_concepts: set = set()
    for idx, row in df.iterrows():
        if row["abstract"] or not q_display_cols:
            continue
        vals = [row[c] for c in q_display_cols if pd.notna(row.get(c))]
        if vals and max(abs(v) for v in vals) >= 1_000:
            scaled_concepts.add(row["concept"])
            for c in q_display_cols:
                if pd.notna(df.at[idx, c]):
                    df.at[idx, c] = df.at[idx, c] / 1_000_000

    # ── LTM / Latest Q ───────────────────────────────────────────────────────
    ltm_col  = COL_LQ if is_balance_sheet else COL_LTM
    ltm_vals = []

    for _, row in df.iterrows():
        if row["abstract"]:
            ltm_vals.append(None)
            continue

        concept = row["concept"]
        scale   = 1_000_000 if concept in scaled_concepts else 1

        if is_balance_sheet:
            ltm_vals.append(
                row[q_display_cols[-1]]
                if q_display_cols and pd.notna(row.get(q_display_cols[-1]))
                else None
            )
            continue

        # Income / CF: prefer proper LTM formula; fall back to 4-Q sum
        if ann_lookup and stub_dates:
            ann_raw = ann_lookup.get(concept)
            if ann_raw is not None:
                stub_raw  = [q_lookup.get(concept, {}).get(d) for d in stub_dates]
                prior_raw = [q_lookup.get(concept, {}).get(p) if p else None
                             for p in prior_dates]

                if all(v is not None for v in stub_raw) and all(v is not None for v in prior_raw):
                    ann_val   = ann_raw / scale
                    stub_sum  = sum(v / scale for v in stub_raw)
                    prior_sum = sum(v / scale for v in prior_raw)
                    ltm_vals.append(ann_val + stub_sum - prior_sum)
                    continue

        # Fallback: sum last 4 consecutive display quarters
        last4       = q_display_cols[-4:] if len(q_display_cols) >= 4 else q_display_cols
        period_vals = [row[c] for c in last4 if pd.notna(row.get(c))]
        ltm_vals.append(sum(period_vals) if len(period_vals) == 4 else None)

    df[ltm_col] = ltm_vals

    # ── QoQ% ─────────────────────────────────────────────────────────────────
    if len(q_display_cols) >= 2:
        q_curr, q_prev = q_display_cols[-1], q_display_cols[-2]
        qoq_vals = []
        for _, row in df.iterrows():
            if row["abstract"]:
                qoq_vals.append(None)
                continue
            c, p = row[q_curr], row[q_prev]
            qoq_vals.append(
                (c - p) / abs(p) if pd.notna(c) and pd.notna(p) and p != 0 else None
            )
        df[COL_QOQ] = qoq_vals

    # ── YoY% ─────────────────────────────────────────────────────────────────
    if len(q_display_cols) >= 5:
        q_curr, q_year_ago = q_display_cols[-1], q_display_cols[-5]
        yoy_vals = []
        for _, row in df.iterrows():
            if row["abstract"]:
                yoy_vals.append(None)
                continue
            c, p = row[q_curr], row[q_year_ago]
            yoy_vals.append(
                (c - p) / abs(p) if pd.notna(c) and pd.notna(p) and p != 0 else None
            )
        df[COL_YOY] = yoy_vals

    return df, q_display_cols, ltm_col


# ── Excel sheet writer ────────────────────────────────────────────────────────

def write_quarterly_sheet(
    wb: Workbook,
    template_stmt,
    q_lookup: dict,
    all_q_dates: list,
    ann_lookup: dict,
    annual_end_date: str | None,
    sheet_title: str,
    is_balance_sheet: bool,
    company_name: str,
    form_type: str,
) -> None:
    df, q_cols, ltm_col = _prepare_quarterly_df(
        template_stmt, q_lookup, all_q_dates, ann_lookup, annual_end_date, is_balance_sheet
    )

    if not q_cols:
        print(f"  [{sheet_title}] no quarterly columns found — sheet skipped", flush=True)
        return

    has_qoq    = COL_QOQ in df.columns
    has_yoy    = COL_YOY in df.columns
    q_labels   = [_q_label(c) for c in q_cols]
    extra_hdrs = [ltm_col] + ([COL_QOQ] if has_qoq else []) + ([COL_YOY] if has_yoy else [])
    min_lvl    = int(df["level"].min())
    n_cols     = 2 + len(q_cols) + len(extra_hdrs)

    ltm_col_idx = 3 + len(q_cols)
    qoq_col_idx = ltm_col_idx + 1
    yoy_col_idx = ltm_col_idx + (2 if has_qoq else 1)

    ws = wb.create_sheet(title=sheet_title[:31])

    # ── Title ─────────────────────────────────────────────────────────────
    title_text = (
        f"{company_name} ({TICKER})  -  {sheet_title}  "
        f"({form_type}, USD Millions)"
    )
    ws.append([title_text] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1)
    tc.font      = Font(bold=True, color=CLR_WHITE, size=11)
    tc.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ── Column headers ────────────────────────────────────────────────────
    ws.append(["Label", "Standard Concept"] + q_labels + extra_hdrs)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HDR_BG)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 2 else "left",
            vertical="center",
            indent=1,
        )
        cell.border = Border(bottom=Side(style="medium", color=CLR_WHITE))
    ws.row_dimensions[2].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    odd = True
    for _, row in df.iterrows():
        is_abstract = bool(row["abstract"])
        indent      = int(row["level"]) - min_lvl
        label       = ("  " * indent) + str(row["label"])
        std_con     = str(row.get("standard_concept", "") or "")
        is_total    = not is_abstract and _is_total_row(str(row["label"]))

        q_vals  = [None if is_abstract else (row[c] if pd.notna(row[c]) else None)
                   for c in q_cols]
        ltm_val = None if is_abstract else (
            row[ltm_col] if pd.notna(row.get(ltm_col)) else None
        )
        qoq_val = None if (is_abstract or not has_qoq) else (
            row[COL_QOQ] if pd.notna(row.get(COL_QOQ)) else None
        )
        yoy_val = None if (is_abstract or not has_yoy) else (
            row[COL_YOY] if pd.notna(row.get(COL_YOY)) else None
        )

        extra_vals = (
            [ltm_val]
            + ([qoq_val] if has_qoq else [])
            + ([yoy_val] if has_yoy else [])
        )
        ws.append([label, std_con] + q_vals + extra_vals)

        data_row = ws.max_row
        fmt = _num_fmt(row, q_cols)

        if is_abstract:
            bg = PatternFill("solid", fgColor=CLR_SECTION_BG)
        elif is_total:
            bg = PatternFill("solid", fgColor=CLR_TOTAL_BG)
        elif odd:
            bg = PatternFill("solid", fgColor=CLR_ALT_BG)
        else:
            bg = None
        if not is_abstract:
            odd = not odd

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(data_row, col_idx)
            if bg:
                cell.fill = bg
            cell.border = Border(bottom=THIN if not is_total else MED)
            cell.font   = Font(bold=(is_abstract or is_total), size=10, color=CLR_DARK)

        ws.cell(data_row, 1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False
        )
        ws.cell(data_row, 2).font      = Font(size=9, color="666666")
        ws.cell(data_row, 2).alignment = Alignment(horizontal="left", vertical="center")

        for col_idx, c in enumerate(q_cols, 3):
            cell = ws.cell(data_row, col_idx)
            if not is_abstract and pd.notna(row[c]):
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ltm_cell = ws.cell(data_row, ltm_col_idx)
        if not is_abstract and ltm_val is not None:
            ltm_cell.number_format = fmt
            ltm_cell.fill = PatternFill("solid", fgColor=CLR_LTM_BG)
        ltm_cell.font      = Font(bold=(not is_abstract), size=10, color=CLR_DARK)
        ltm_cell.alignment = Alignment(horizontal="right", vertical="center")

        if has_qoq:
            qoq_cell = ws.cell(data_row, qoq_col_idx)
            if not is_abstract and qoq_val is not None:
                qoq_cell.number_format = PCT_FMT
            qoq_cell.font      = Font(size=10, color=CLR_DARK)
            qoq_cell.alignment = Alignment(horizontal="right", vertical="center")

        if has_yoy:
            yoy_cell = ws.cell(data_row, yoy_col_idx)
            if not is_abstract and yoy_val is not None:
                yoy_cell.number_format = PCT_FMT
            yoy_cell.font      = Font(size=10, color=CLR_DARK)
            yoy_cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[data_row].height = 15

    # ── Column widths + freeze ────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 52
    ws.column_dimensions[get_column_letter(2)].width = 26
    for i in range(len(q_cols)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12
    ws.column_dimensions[get_column_letter(ltm_col_idx)].width = 12
    if has_qoq:
        ws.column_dimensions[get_column_letter(qoq_col_idx)].width = 9
    if has_yoy:
        ws.column_dimensions[get_column_letter(yoy_col_idx)].width = 9
    ws.freeze_panes = ws.cell(3, 3)


# ── Main ──────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

company = Company(TICKER)
step(f"Looked up {company.name} ({TICKER})  CIK {company.cik}", t_start)

# Detect quarterly form type: 10-Q domestic, 6-K foreign
form_type = None
filings   = None
for form in ("10-Q", "6-K"):
    candidate = company.get_filings(form=form, amendments=False).head(N_FILINGS)
    if len(candidate) > 0:
        form_type = form
        filings   = candidate
        break

if not filings:
    raise SystemExit(f"No 10-Q or 6-K filings found for {TICKER}.")

step(
    f"Found {len(filings)} {form_type} filing(s) "
    f"({filings[len(filings)-1].filing_date} to {filings[0].filing_date})",
    t_start,
)

# Annual base for LTM (latest 10-K or 20-F)
step("Getting latest annual filing for LTM base...", t_start)
annual_financials = company.get_financials()

# Template: latest quarterly filing provides row order and section structure
step("Getting quarterly template structure...", t_start)
template_financials = company.get_quarterly_financials()
if template_financials is None:
    raise SystemExit(
        f"Could not parse financials from latest {form_type} filing for {TICKER}. "
        "Most 6-K filings are press releases without XBRL. "
        "Quarterly analysis requires XBRL-tagged 10-Q filings (US domestic companies)."
    )

STMT_CONFIGS = {
    "Income Statement": ("income_statement",   False),
    "Balance Sheet":    ("balance_sheet",      True),
    "Cash Flow":        ("cashflow_statement", False),
}

# Build quarterly lookup — one pass, all filings, all statements
step(f"Extracting data from {len(filings)} individual {form_type} filings...", t_start)
q_lookups    = {name: {} for name in STMT_CONFIGS}
q_all_dates  = {name: [] for name in STMT_CONFIGS}
# CF-specific: 10-Q cash flow is YTD for Q2/Q3; collect raw YTD values then derive standalones
cf_ytd_lookup: dict = {}  # concept -> {date_str: raw_ytd_value}
cf_ytd_dates:  list = []  # all CF dates collected from 10-Qs (Q1 standalone or Q2/Q3 YTD)
n_parsed = 0

for filing in filings:
    try:
        fin = Financials.extract(filing)
        if fin is None:
            continue
        n_parsed += 1
        for name, (method, is_bs) in STMT_CONFIGS.items():
            try:
                stmt = getattr(fin, method)()
                if stmt is None:
                    continue
                df_raw = stmt.to_dataframe(view="standard")

                if name == "Cash Flow":
                    # Q1 filings label CF as (Q1); Q2/Q3 label the YTD as (YTD).
                    # Collect all into cf_ytd_lookup; standalone values computed below.
                    primary_col = _q_col(df_raw) or _ytd_col(df_raw)
                    if primary_col is None:
                        continue
                    date_str = _end_date(primary_col)
                    if not date_str or date_str in cf_ytd_dates:
                        continue
                    cf_ytd_dates.append(date_str)
                    for concept, val in _extract_values(df_raw, primary_col).items():
                        cf_ytd_lookup.setdefault(concept, {})[date_str] = val
                else:
                    primary_col = _instant_col(df_raw) if is_bs else _q_col(df_raw)
                    if primary_col is None:
                        continue
                    date_str = _end_date(primary_col)
                    if not date_str or date_str in q_all_dates[name]:
                        continue
                    q_all_dates[name].append(date_str)
                    for concept, val in _extract_values(df_raw, primary_col).items():
                        q_lookups[name].setdefault(concept, {})[date_str] = val
            except Exception:
                pass
    except Exception:
        pass

step(f"Parsed {n_parsed}/{len(filings)} filings successfully", t_start)

# ── Convert CF YTD to standalone quarterly values ────────────────────────────
# Q1 (3-month) YTD = standalone.  Q2 = YTD[Q2] - YTD[Q1].  Q3 = YTD[Q3] - YTD[Q2].
# A date gap <= 130 days between consecutive YTD dates = same fiscal year.
cf_sorted = sorted(cf_ytd_dates)
for i, date in enumerate(cf_sorted):
    prev_date = cf_sorted[i - 1] if i > 0 else None
    same_fy   = False
    if prev_date:
        dt      = datetime.strptime(date,      "%Y-%m-%d")
        prev_dt = datetime.strptime(prev_date, "%Y-%m-%d")
        same_fy = (dt - prev_dt).days <= 130
    q_all_dates["Cash Flow"].append(date)
    for concept, ytd_vals in cf_ytd_lookup.items():
        curr = ytd_vals.get(date)
        if curr is None:
            continue
        if same_fy:
            prev = ytd_vals.get(prev_date)
            val  = curr - prev if prev is not None else curr
        else:
            val = curr  # Q1 or first in FY: YTD == standalone
        q_lookups["Cash Flow"].setdefault(concept, {})[date] = val

for name in STMT_CONFIGS:
    q_all_dates[name] = sorted(q_all_dates[name])

if not any(q_all_dates[n] for n in STMT_CONFIGS):
    raise SystemExit(
        f"No XBRL quarterly data found in {form_type} filings for {TICKER}. "
        "Most 6-K filings are press releases without XBRL attachments. "
        "Try a US domestic ticker such as AAPL, MSFT, AMZN, TSLA, or GOOG."
    )

# Build annual lookup for LTM
step("Extracting annual data for LTM base...", t_start)
ann_lookups   = {}
ann_end_dates = {}

if annual_financials:
    for name, (method, is_bs) in STMT_CONFIGS.items():
        try:
            stmt = getattr(annual_financials, method)()
            df_raw      = stmt.to_dataframe(view="standard")
            primary_col = _instant_col(df_raw) if is_bs else _fy_col(df_raw)
            if primary_col:
                ann_end_dates[name] = _end_date(primary_col)
                ann_lookups[name]   = _extract_values(df_raw, primary_col)
            else:
                ann_end_dates[name] = None
                ann_lookups[name]   = {}
        except Exception:
            ann_end_dates[name] = None
            ann_lookups[name]   = {}
    step(
        f"Annual base: {ann_end_dates.get('Income Statement', 'n/a')} "
        f"({len(ann_lookups.get('Income Statement', {}))} IS concepts)",
        t_start,
    )
else:
    step("No annual financials found — LTM will use 4-quarter fallback", t_start)
    for name in STMT_CONFIGS:
        ann_end_dates[name] = None
        ann_lookups[name]   = {}

# Console preview — income statement
df_prev, qc_prev, _ = _prepare_quarterly_df(
    template_financials.income_statement(),
    q_lookups["Income Statement"],
    q_all_dates["Income Statement"],
    ann_lookups.get("Income Statement", {}),
    ann_end_dates.get("Income Statement"),
    False,
)
preview_q    = qc_prev[-4:]
preview_cols = preview_q + ([COL_LTM] if COL_LTM in df_prev.columns else [])
print(f"\n--- Income Statement (recent quarters, {form_type}, USD Millions) ---", flush=True)
preview = df_prev[["label"] + preview_cols].copy()
preview.columns = (
    ["Label"]
    + [_q_label(c) for c in preview_q]
    + (["LTM"] if COL_LTM in df_prev.columns else [])
)
preview["Label"] = preview["Label"].str.strip()
for c in preview.columns[1:]:
    preview[c] = preview[c].map(lambda x: f"{x:,.1f}" if pd.notna(x) else "")
print(preview.to_string(index=False), flush=True)

# Excel export
out_dir  = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{TICKER}_quarterly.xlsx"

step(f"Writing Excel workbook to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)

for sheet_name, (_, is_bs) in STMT_CONFIGS.items():
    write_quarterly_sheet(
        wb,
        getattr(template_financials, STMT_CONFIGS[sheet_name][0])(),
        q_lookups[sheet_name],
        q_all_dates[sheet_name],
        ann_lookups.get(sheet_name, {}),
        ann_end_dates.get(sheet_name),
        sheet_name,
        is_bs,
        company.name,
        form_type,
    )

if not wb.sheetnames:
    raise SystemExit("No sheets were written — no quarterly data found.")

wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel - {len(wb.sheetnames)} tab(s).", t_start)
