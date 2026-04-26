r"""
Pull N years of annual financial statements from SEC EDGAR via XBRLS and
export a single Excel workbook with one sheet per statement.

Designed for financial modelling:
- Structure (row order, section headers, indentation) from the latest single
  filing — matches what you see in the actual 10-K / 20-F document
- Values across all years from XBRLS multi-period stitching
- Values scaled to USD millions (except per-share metrics)
- standard_concept column included for cross-company mapping
- All-NaN rows dropped (concepts not reported by this company)
- Automatically tries 10-K then 20-F (for foreign private issuers)

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\98_multiperiod_financials.py
"""
import os
import re
import time
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
os.environ["EDGAR_IDENTITY"] = os.getenv("SEC_USER_AGENT", "")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from edgar import Company
from edgar.xbrl import XBRLS

# ── Configuration ────────────────────────────────────────────────────────────
TICKER = "SPOT"
YEARS  = 5      # number of annual filings to stitch (older filings may lack XBRL)

# ── Colour palette ───────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"
CLR_HDR_BG     = "4472C4"
CLR_SECTION_BG = "D9E1F2"   # light blue  — section header rows
CLR_TOTAL_BG   = "EBF0FA"   # pale blue   — subtotal / total rows
CLR_ALT_BG     = "F9F9F9"   # off-white   — alternating data rows
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"
THIN           = Side(style="thin",   color="CCCCCC")
MED            = Side(style="medium", color="AAAAAA")

TOTAL_KEYWORDS = (
    "total", "gross margin", "operating income", "net income",
    "earnings before", "ebitda", "free cash flow", "gross profit",
)


def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


def _period_cols(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if re.match(r"\d{4}-\d{2}-\d{2}$", c)]


def _fy_label(date_str: str) -> str:
    return f"FY{date_str[:4]}"


def _is_total_row(label: str) -> bool:
    low = label.strip().lower()
    return any(kw in low for kw in TOTAL_KEYWORDS)


def _prepare_structured_df(
    template_stmt,
    stitched_stmt,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    """
    Merge single-filing structure (row order, levels, section headers) with
    XBRLS multi-period values joined by XBRL concept.

    Returns (df, period_cols, fy_labels).
    df columns: label, level, abstract, standard_concept, <date>, <date>, ...
    Values scaled to millions (per-share rows left unchanged).
    """
    # ── Template: structure only ──────────────────────────────────────────
    tmpl = template_stmt.to_dataframe(view="standard")

    # ── Stitched: multi-period values ─────────────────────────────────────
    stitched = stitched_stmt.to_dataframe()
    period_cols = _period_cols(stitched)

    for c in period_cols:
        stitched[c] = pd.to_numeric(stitched[c], errors="coerce")

    # Build concept → values lookup (first non-dimensional match wins)
    stitch_lookup: dict[str, dict] = {}
    for _, srow in stitched.iterrows():
        concept = srow["concept"]
        if concept not in stitch_lookup:
            stitch_lookup[concept] = {c: srow[c] for c in period_cols}

    # ── Merge: walk template rows in filing order ─────────────────────────
    rows = []
    for _, trow in tmpl.iterrows():
        if trow.get("dimension", False):
            continue   # skip per-dimension breakdowns (not aligned across years)

        concept     = trow["concept"]
        sc          = trow.get("standard_concept")
        is_abstract = bool(trow.get("abstract", False))
        level       = int(trow.get("level", 1))

        row_data = {
            "label":            trow["label"],
            "level":            level,
            "abstract":         is_abstract,
            "standard_concept": sc if pd.notna(sc) else "",
        }

        if is_abstract:
            for c in period_cols:
                row_data[c] = None
        else:
            hist = stitch_lookup.get(concept, {})
            for c in period_cols:
                row_data[c] = hist.get(c)

        rows.append(row_data)

    df = pd.DataFrame(rows)

    # Drop rows with no values across all periods (unmapped concepts)
    data_mask = ~df["abstract"]
    df = df[df["abstract"] | df.loc[data_mask, period_cols].notna().any(axis=1)]
    df = df.reset_index(drop=True)

    # ── Scale to millions ─────────────────────────────────────────────────
    for idx, row in df.iterrows():
        if row["abstract"]:
            continue
        vals = [row[c] for c in period_cols if pd.notna(row[c])]
        if vals and max(abs(v) for v in vals) >= 1_000:
            for c in period_cols:
                if pd.notna(df.at[idx, c]):
                    df.at[idx, c] = df.at[idx, c] / 1_000_000

    fy_labels = [_fy_label(c) for c in period_cols]
    return df, period_cols, fy_labels


def _num_fmt(row: pd.Series, period_cols: list[str]) -> str:
    vals = [row[c] for c in period_cols if pd.notna(row[c])]
    if not vals:
        return "#,##0"
    return "#,##0.00" if max(abs(v) for v in vals) < 1_000 else "#,##0"


def write_statement_sheet(
    wb: Workbook,
    template_stmt,
    stitched_stmt,
    sheet_title: str,
    company_name: str,
    form_type: str,
) -> None:
    df, period_cols, fy_labels = _prepare_structured_df(template_stmt, stitched_stmt)
    min_lvl      = int(df["level"].min())
    n_value_cols = len(period_cols)
    n_cols       = 2 + n_value_cols   # label | standard_concept | FY... | FY...

    ws = wb.create_sheet(title=sheet_title[:31])

    # ── Title row ─────────────────────────────────────────────────────────
    title_text = (
        f"{company_name} ({TICKER})  -  {sheet_title}  "
        f"({form_type}, USD Millions)"
    )
    ws.append([title_text] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1)
    tc.font      = Font(bold=True, color=CLR_WHITE, size=11)
    tc.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ── Column header row ─────────────────────────────────────────────────
    ws.append(["Label", "Standard Concept"] + fy_labels)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HDR_BG)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 2 else "left",
            vertical="center",
            indent=1,
        )
        cell.border = Border(bottom=Side(style="medium", color=CLR_WHITE))
    ws.row_dimensions[2].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    odd = True   # for alternating row colour on non-abstract rows
    for _, row in df.iterrows():
        is_abstract = bool(row["abstract"])
        indent      = int(row["level"]) - min_lvl
        label       = ("  " * indent) + str(row["label"])
        std_con     = str(row["standard_concept"])
        values      = [None if is_abstract else (row[c] if pd.notna(row[c]) else None)
                       for c in period_cols]
        ws.append([label, std_con] + values)

        data_row = ws.max_row
        is_total = not is_abstract and _is_total_row(str(row["label"]))
        fmt      = _num_fmt(row, period_cols)

        # Row background
        if is_abstract:
            bg = PatternFill("solid", fgColor=CLR_SECTION_BG)
        elif is_total:
            bg = PatternFill("solid", fgColor=CLR_TOTAL_BG)
        elif odd:
            bg = PatternFill("solid", fgColor=CLR_ALT_BG)
        else:
            bg = None
        if not is_abstract:
            odd = not odd

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(data_row, col_idx)
            if bg:
                cell.fill = bg
            cell.border = Border(bottom=THIN if not is_total else MED)
            cell.font   = Font(bold=(is_abstract or is_total), size=10, color=CLR_DARK)

        # Label cell alignment
        ws.cell(data_row, 1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False
        )
        # standard_concept cell
        ws.cell(data_row, 2).font      = Font(size=9, color="666666")
        ws.cell(data_row, 2).alignment = Alignment(horizontal="left", vertical="center")

        # Value cells
        for col_idx, c in enumerate(period_cols, 3):
            cell = ws.cell(data_row, col_idx)
            if not is_abstract and pd.notna(row[c]):
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[data_row].height = 15

    # ── Column widths + freeze ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26
    for i in range(n_value_cols):
        ws.column_dimensions[chr(ord("C") + i)].width = 14
    ws.freeze_panes = ws.cell(3, 3)


# ── Main ─────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

company = Company(TICKER)
step(f"Looked up {company.name} ({TICKER})  CIK {company.cik}", t_start)

# Detect form type
form_type = None
filings   = None
for form in ("10-K", "20-F"):
    candidate = company.get_filings(form=form).head(YEARS)
    if len(candidate) > 0:
        form_type = form
        filings   = candidate
        break

if not filings:
    raise SystemExit(f"No 10-K or 20-F filings found for {TICKER}.")

step(
    f"Found {len(filings)} {form_type} filing(s) "
    f"({filings[len(filings)-1].filing_date} to {filings[0].filing_date})",
    t_start,
)

# Template: latest single filing — provides filing-order structure
step("Getting latest filing structure (template)...", t_start)
template_financials = company.get_financials()
if template_financials is None:
    raise SystemExit("Could not parse financials from latest filing.")

# Stitched: multi-period values
step("Stitching XBRL across filings — may take 30-90s the first time...", t_start)
xbrls = XBRLS.from_filings(filings)
step("XBRLS ready", t_start)

STMT_METHODS = {
    "Income Statement": "income_statement",
    "Balance Sheet":    "balance_sheet",
    "Cash Flow":        "cashflow_statement",
}

step("Extracting and merging statements...", t_start)
template_stmts = {
    name: getattr(template_financials, method)()
    for name, method in STMT_METHODS.items()
}
stitched_stmts = {
    name: getattr(xbrls.statements, method)()
    for name, method in STMT_METHODS.items()
}

# Console preview (income statement)
df_prev, pc_prev, fy_prev = _prepare_structured_df(
    template_stmts["Income Statement"],
    stitched_stmts["Income Statement"],
)
print(f"\n--- Income Statement ({form_type}, USD Millions) ---", flush=True)
preview = df_prev[["label"] + pc_prev].copy()
preview.columns = ["Label"] + fy_prev
preview["Label"] = preview["Label"].str.strip()
preview[fy_prev] = preview[fy_prev].map(
    lambda x: f"{x:,.1f}" if pd.notna(x) else ""
)
print(preview.to_string(index=False), flush=True)

# Excel export
out_dir  = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{TICKER}_multiperiod_financials.xlsx"

step(f"Writing Excel workbook to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)

for sheet_name in STMT_METHODS:
    write_statement_sheet(
        wb,
        template_stmts[sheet_name],
        stitched_stmts[sheet_name],
        sheet_name,
        company.name,
        form_type,
    )

wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel - three tabs.", t_start)
