r"""
Pull Apple's three core financial statements from the latest 10-K and:
  1. Print all three to the console
  2. Export to a single Excel workbook with three sheets
     (Income Statement, Balance Sheet, Cash Flow)

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\99_income_statement.py

Notes:
- The first run can take 30-90s while edgartools downloads the latest 10-K
  XBRL filing (~10MB) and parses it. Subsequent runs are cached and fast.
- All print() calls use flush=True so you see progress in real time
  rather than thinking the script has hung.
"""
import os
import re
import time
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()
os.environ["EDGAR_IDENTITY"] = os.getenv("SEC_USER_AGENT", "")

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from edgar import Company


def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


# ── Colour palette ──────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"   # dark blue  — sheet title row
CLR_HEADER_BG  = "4472C4"   # medium blue — period header row
CLR_SECTION_BG = "D9E1F2"   # light blue  — abstract / section rows
CLR_ALT_BG     = "F2F2F2"   # very light grey — indent-1 data rows
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"

THIN = Side(style="thin", color="CCCCCC")


def _period_cols(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if re.match(r"\d{4}-\d{2}-\d{2}", c)]


def _num_fmt(row: pd.Series, period_cols: list[str]) -> str:
    """Use 2 d.p. for EPS-scale numbers (abs max < 1 000), else integer."""
    vals = [row[c] for c in period_cols if pd.notna(row[c])]
    if not vals:
        return "#,##0"
    return "#,##0.00" if max(abs(v) for v in vals) < 1_000 else "#,##0"


def write_statement_sheet(wb: Workbook, stmt, sheet_title: str) -> None:
    df = stmt.to_dataframe(view="standard")
    periods = sorted(_period_cols(df))
    min_lvl = int(df["level"].min())
    n_cols = 2 + len(periods)   # label | standard_concept | period...

    # Coerce and scale to USD millions (skip per-share rows: max(abs) < 1 000)
    for c in periods:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    for idx, row in df.iterrows():
        if row["abstract"]:
            continue
        vals = [row[c] for c in periods if pd.notna(row[c])]
        if vals and max(abs(v) for v in vals) >= 1_000:
            for c in periods:
                if pd.notna(df.at[idx, c]):
                    df.at[idx, c] = df.at[idx, c] / 1_000_000

    ws = wb.create_sheet(title=sheet_title[:31])

    # ── Title row ──────────────────────────────────────────────────────────
    title_text = f"{sheet_title}  (USD Millions)"
    ws.append([title_text] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(1, 1)
    title_cell.font  = Font(bold=True, color=CLR_WHITE, size=12)
    title_cell.fill  = PatternFill("solid", fgColor=CLR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 20

    # ── Period header row ──────────────────────────────────────────────────
    ws.append(["Label", "Standard Concept"] + periods)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.font  = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill  = PatternFill("solid", fgColor=CLR_HEADER_BG)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 2 else "left",
            vertical="center",
            indent=1,
        )
        cell.border = Border(bottom=Side(style="medium", color=CLR_WHITE))
    ws.row_dimensions[2].height = 18

    # ── Data rows ──────────────────────────────────────────────────────────
    for _, row in df.iterrows():
        indent       = int(row["level"]) - min_lvl
        label        = ("  " * indent) + str(row["label"])
        std_con      = str(row.get("standard_concept", "") or "")
        is_abstract  = bool(row["abstract"])
        is_dimension = bool(row["dimension"])
        values       = [None if is_abstract else (row[c] if pd.notna(row[c]) else None)
                        for c in periods]
        ws.append([label, std_con] + values)

        data_row = ws.max_row
        fmt = _num_fmt(row, periods)

        # Label cell style
        lbl_cell = ws.cell(data_row, 1)
        lbl_cell.font = Font(
            bold=is_abstract,
            italic=is_dimension,
            size=10,
            color=CLR_DARK,
        )
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Standard concept cell
        sc_cell = ws.cell(data_row, 2)
        sc_cell.font      = Font(size=9, color="666666")
        sc_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Row background
        if is_abstract:
            bg = PatternFill("solid", fgColor=CLR_SECTION_BG)
        elif indent == 0:
            bg = PatternFill("solid", fgColor=CLR_ALT_BG)
        else:
            bg = None

        if bg:
            for c in range(1, n_cols + 1):
                ws.cell(data_row, c).fill = bg

        # Value cell style
        for col_idx, c in enumerate(periods, 3):
            cell = ws.cell(data_row, col_idx)
            if not is_abstract and pd.notna(row[c]):
                cell.number_format = fmt
            cell.font      = Font(bold=is_abstract, size=10, color=CLR_DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border    = Border(bottom=THIN)

        ws.row_dimensions[data_row].height = 15

    # ── Column widths + freeze ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 26
    for i in range(len(periods)):
        ws.column_dimensions[chr(ord("C") + i)].width = 20
    ws.freeze_panes = ws.cell(3, 3)   # freeze title + header rows, label + concept cols


# ── Main ────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

pd.set_option("display.max_rows", 80)
pd.set_option("display.width", 180)
pd.set_option("display.float_format", lambda x: f"{x:,.0f}")

TICKER = "TSLA"
company = Company(TICKER)
step(f"Looked up {company.name} ({TICKER}) CIK {company.cik}", t_start)

step("Fetching latest 10-K and parsing XBRL (this can take 30-90s the first time)...", t_start)
financials = company.get_financials()
step("Got financials object", t_start)

if financials is None:
    raise SystemExit("No financials parsed - likely no recent 10-K available.")

step("Extracting Income Statement...", t_start)
income = financials.income_statement()
step("Extracting Balance Sheet...", t_start)
balance = financials.balance_sheet()
step("Extracting Cash Flow Statement...", t_start)
cashflow = financials.cashflow_statement()
step("Extracting Comprehensive Income...", t_start)
comp_income = financials.comprehensive_income()

statements = {
    "Income Statement":    income,
    "Balance Sheet":       balance,
    "Cash Flow":           cashflow,
    "Comprehensive Income": comp_income,
}

# --- Print each to console ---
for name, stmt in statements.items():
    print(f"\n--- {name} (latest 10-K, $) ---", flush=True)
    print(stmt, flush=True)

# --- Export all three to a single styled Excel workbook ---
out_dir = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{TICKER}_financials.xlsx"

step(f"Writing Excel workbook to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)   # remove the default empty sheet

for sheet_name, stmt in statements.items():
    write_statement_sheet(wb, stmt, sheet_name)

wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel - three tabs.", t_start)
