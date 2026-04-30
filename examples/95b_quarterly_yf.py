r"""
95b_quarterly_yf.py — Quarterly financials via yfinance (foreign-issuer fallback)

Same purpose as 95_quarterly.py but sources data from yfinance instead of
EDGAR XBRL. Use this for foreign issuers that file 6-K without XBRL
quarterly data (ASML, TSM, SPOT) or any ticker requiring a quick
multi-quarter view without pulling full XBRL filings.

Key difference vs 95_quarterly.py:
  Source: yfinance (RESTATED figures) not point-in-time SEC filings.
  A quarter that was later restated shows the revised figure, not what
  was reported live. Fine for modelling; wrong for backtesting.

Output per statement:
  N quarterly columns (oldest left, newest right)
  LTM  : sum of last 4 quarters (IS and CF only)
  Latest Q: most recent quarter (Balance Sheet only)
  QoQ% : latest Q vs prior Q
  YoY% : latest Q vs same Q one year ago (requires >= 5 quarters)

Currency note: values are in the company's reporting currency
  (shown in the sheet title row).  TSM -> TWD, ASML -> EUR, SPOT/US -> USD.

Run from the repo root:
    cd E:\github_awesome\fundamental\edgartools
    uv run python examples\95b_quarterly_yf.py
"""
import time
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
TICKER     = "ASML"
N_QUARTERS = 6   # quarterly periods to display (yfinance provides ~6)

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_TITLE_BG   = "2F5496"
CLR_HDR_BG     = "4472C4"
CLR_SECTION_BG = "D9E1F2"
CLR_TOTAL_BG   = "EBF0FA"
CLR_LTM_BG     = "E2EFDA"
CLR_ALT_BG     = "F9F9F9"
CLR_WHITE      = "FFFFFF"
CLR_DARK       = "1F1F1F"
THIN           = Side(style="thin",   color="CCCCCC")
MED            = Side(style="medium", color="AAAAAA")

COL_LTM = "LTM"
COL_LQ  = "Latest Q"
COL_QOQ = "QoQ %"
COL_YOY = "YoY %"
PCT_FMT = "+0.0%;-0.0%;--"

# ── Statement row specs ───────────────────────────────────────────────────────
# {"section": str}                        -> section header (no data)
# {"label": str, "key": str, "total": bool} -> data row; key must match
#                                              a yfinance DataFrame index label.
#                                              Missing keys are silently skipped.

IS_SPEC = [
    {"section": "Revenue"},
    {"label": "Total Revenue",         "key": "Total Revenue",                      "total": True},
    {"label": "Cost of Revenue",       "key": "Cost Of Revenue",                    "total": False},
    {"label": "Gross Profit",          "key": "Gross Profit",                       "total": True},
    {"section": "Operating Expenses"},
    {"label": "R&D",                   "key": "Research And Development",           "total": False},
    {"label": "SG&A",                  "key": "Selling General And Administration", "total": False},
    {"label": "Operating Expense",     "key": "Operating Expense",                 "total": False},
    {"label": "Operating Income",      "key": "Operating Income",                  "total": True},
    {"section": "Non-Operating"},
    {"label": "Net Interest Income",   "key": "Net Interest Income",               "total": False},
    {"label": "Pretax Income",         "key": "Pretax Income",                     "total": True},
    {"label": "Tax Provision",         "key": "Tax Provision",                     "total": False},
    {"label": "Net Income",            "key": "Net Income",                        "total": True},
    {"section": "Supplemental"},
    {"label": "EBITDA",                "key": "EBITDA",                            "total": True},
    {"section": "Per Share"},
    {"label": "Basic EPS",             "key": "Basic EPS",                         "total": False},
    {"label": "Diluted EPS",           "key": "Diluted EPS",                       "total": False},
    # is_avg=True: LTM uses latest-Q (not sum); weighted-avg shares don't sum across quarters
    {"label": "Basic Shares",          "key": "Basic Average Shares",              "total": False, "is_avg": True},
    {"label": "Diluted Shares",        "key": "Diluted Average Shares",            "total": False, "is_avg": True},
]

BS_SPEC = [
    {"section": "Assets"},
    {"label": "Cash & Equivalents",    "key": "Cash And Cash Equivalents",              "total": False},
    {"label": "Current Assets",        "key": "Current Assets",                         "total": True},
    {"label": "Net PPE",               "key": "Net PPE",                                "total": False},
    {"label": "Total Assets",          "key": "Total Assets",                           "total": True},
    {"section": "Liabilities"},
    {"label": "Current Liabilities",   "key": "Current Liabilities",                   "total": True},
    {"label": "Long Term Debt",        "key": "Long Term Debt",                        "total": False},
    {"label": "Total Debt",            "key": "Total Debt",                            "total": False},
    {"label": "Total Liabilities",     "key": "Total Liabilities Net Minority Interest","total": True},
    {"section": "Equity"},
    {"label": "Stockholders Equity",   "key": "Stockholders Equity",                   "total": True},
    {"label": "Retained Earnings",     "key": "Retained Earnings",                     "total": False},
    {"section": "Metrics"},
    {"label": "Working Capital",       "key": "Working Capital",                       "total": False},
    {"label": "Invested Capital",      "key": "Invested Capital",                      "total": False},
]

CF_SPEC = [
    {"section": "Operating"},
    {"label": "Operating Cash Flow",   "key": "Operating Cash Flow",                  "total": True},
    {"label": "D&A",                   "key": "Depreciation And Amortization",        "total": False},
    {"label": "Change in Working Cap", "key": "Change In Working Capital",            "total": False},
    {"section": "Investing"},
    {"label": "Capital Expenditure",   "key": "Capital Expenditure",                  "total": False},
    {"label": "Investing Cash Flow",   "key": "Investing Cash Flow",                  "total": False},
    {"section": "Financing"},
    {"label": "Cash Dividends",        "key": "Cash Dividends Paid",                  "total": False},
    {"label": "Buybacks",              "key": "Repurchase Of Capital Stock",          "total": False},
    {"label": "Financing Cash Flow",   "key": "Financing Cash Flow",                  "total": False},
    {"section": "Summary"},
    {"label": "Net Change in Cash",    "key": "Changes In Cash",                      "total": False},
    {"label": "Free Cash Flow",        "key": "Free Cash Flow",                       "total": True},
]


# ── Utilities ─────────────────────────────────────────────────────────────────

def step(msg: str, t0: float) -> float:
    print(f"[{time.time() - t0:6.1f}s] {msg}", flush=True)
    return time.time()


def _q_label(ts: pd.Timestamp) -> str:
    return ts.strftime("%b %Y")


def _safe(df: pd.DataFrame, key: str, col: pd.Timestamp):
    """Numeric value from df.loc[key, col], or None."""
    if key not in df.index:
        return None
    try:
        val = pd.to_numeric(df.loc[key, col], errors="coerce")
        return val if pd.notna(val) else None
    except Exception:
        return None


def _select_cols(df: pd.DataFrame) -> list:
    """N_QUARTERS most recent columns, sorted oldest first."""
    return sorted(df.columns)[-N_QUARTERS:]


# ── Core: build row list for one statement ────────────────────────────────────

def _build_rows(
    spec: list,
    yf_df: pd.DataFrame,
    q_cols: list,
    is_balance_sheet: bool,
) -> list[dict]:
    """
    Returns a flat list of row dicts ready for Excel rendering:
      {"is_header": True,  "label": str}
      {"is_header": False, "label": str, "is_total": bool, "scale": int,
       "vals": {ts: val|None}, "ltm": val|None, "qoq": pct|None, "yoy": pct|None}

    scale=1 means per-unit (EPS, shares); scale=1_000_000 means scaled to millions.
    Empty sections (all data rows skipped) are suppressed.
    """
    rows = []
    for entry in spec:
        if "section" in entry:
            rows.append({"is_header": True, "label": entry["section"]})
            continue

        key = entry["key"]
        if key not in yf_df.index:
            continue

        raw = {col: _safe(yf_df, key, col) for col in q_cols}

        vals_exist = [v for v in raw.values() if v is not None]
        scale = 1_000_000 if (vals_exist and max(abs(v) for v in vals_exist) >= 1_000) else 1
        scaled = {col: (v / scale if v is not None else None) for col, v in raw.items()}

        # LTM or Latest Q
        last4   = q_cols[-4:] if len(q_cols) >= 4 else q_cols
        is_avg  = entry.get("is_avg", False)   # True for shares — use latest Q, not sum
        if is_balance_sheet or is_avg:
            ltm = scaled.get(q_cols[-1])
        else:
            vals4 = [scaled[c] for c in last4 if scaled.get(c) is not None]
            ltm   = sum(vals4) if len(vals4) == len(last4) else None

        # QoQ
        qoq = None
        if len(q_cols) >= 2:
            c, p = scaled.get(q_cols[-1]), scaled.get(q_cols[-2])
            if c is not None and p is not None and p != 0:
                qoq = (c - p) / abs(p)

        # YoY (same quarter one year ago = 4 periods back, index [-5] in display list)
        yoy = None
        if len(q_cols) >= 5:
            c, p = scaled.get(q_cols[-1]), scaled.get(q_cols[-5])
            if c is not None and p is not None and p != 0:
                yoy = (c - p) / abs(p)

        rows.append({
            "is_header": False,
            "label":     entry["label"],
            "is_total":  entry.get("total", False),
            "scale":     scale,
            "vals":      scaled,
            "ltm":       ltm,
            "qoq":       qoq,
            "yoy":       yoy,
        })

    # Suppress empty sections
    out, pending_hdr = [], None
    for row in rows:
        if row["is_header"]:
            pending_hdr = row
        else:
            if pending_hdr:
                out.append(pending_hdr)
                pending_hdr = None
            out.append(row)
    return out


# ── Excel sheet writer ────────────────────────────────────────────────────────

def write_sheet(
    wb: Workbook,
    spec: list,
    yf_df: pd.DataFrame,
    q_cols: list,
    is_balance_sheet: bool,
    sheet_title: str,
    company_name: str,
    currency: str,
) -> None:
    rows = _build_rows(spec, yf_df, q_cols, is_balance_sheet)
    if not any(not r["is_header"] for r in rows):
        print(f"  [{sheet_title}] no data rows found - sheet skipped", flush=True)
        return

    has_qoq    = len(q_cols) >= 2
    has_yoy    = len(q_cols) >= 5
    ltm_label  = COL_LQ if is_balance_sheet else COL_LTM
    extra_hdrs = [ltm_label] + ([COL_QOQ] if has_qoq else []) + ([COL_YOY] if has_yoy else [])
    q_labels   = [_q_label(c) for c in q_cols]
    n_cols     = 1 + len(q_cols) + len(extra_hdrs)

    ltm_col_idx = 1 + len(q_cols) + 1          # 1-indexed
    qoq_col_idx = ltm_col_idx + 1
    yoy_col_idx = ltm_col_idx + (2 if has_qoq else 1)

    ws = wb.create_sheet(title=sheet_title[:31])

    # Title
    title = (
        f"{company_name} ({TICKER})  -  {sheet_title}  "
        f"({currency} Millions  |  source: yfinance - restated figures)"
    )
    ws.append([title] + [""] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1)
    tc.font      = Font(bold=True, color=CLR_WHITE, size=11)
    tc.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # Column headers
    ws.append(["Label"] + q_labels + extra_hdrs)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(2, col_idx)
        cell.font      = Font(bold=True, color=CLR_WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HDR_BG)
        cell.alignment = Alignment(
            horizontal="right" if col_idx > 1 else "left",
            vertical="center", indent=1,
        )
        cell.border = Border(bottom=Side(style="medium", color=CLR_WHITE))
    ws.row_dimensions[2].height = 18

    # Data rows
    odd = True
    for row in rows:
        if row["is_header"]:
            ws.append([row["label"]] + [""] * (n_cols - 1))
            hr = ws.max_row
            for col in range(1, n_cols + 1):
                cell = ws.cell(hr, col)
                cell.fill      = PatternFill("solid", fgColor=CLR_SECTION_BG)
                cell.border    = Border(bottom=THIN)
                cell.font      = Font(bold=True, size=10, color=CLR_DARK)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[hr].height = 16
            continue

        q_vals  = [row["vals"].get(c) for c in q_cols]
        ltm_val = row["ltm"]
        qoq_val = row["qoq"] if has_qoq else None
        yoy_val = row["yoy"] if has_yoy else None
        ws.append(
            [row["label"]] + q_vals + [ltm_val]
            + ([qoq_val] if has_qoq else []) + ([yoy_val] if has_yoy else [])
        )
        dr       = ws.max_row
        is_total = row["is_total"]
        fmt      = "#,##0.00" if row["scale"] == 1 else "#,##0"

        if is_total:
            bg = PatternFill("solid", fgColor=CLR_TOTAL_BG)
        elif odd:
            bg = PatternFill("solid", fgColor=CLR_ALT_BG)
        else:
            bg = None
        odd = not odd

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(dr, col_idx)
            if bg:
                cell.fill = bg
            cell.border = Border(bottom=THIN if not is_total else MED)
            cell.font   = Font(bold=is_total, size=10, color=CLR_DARK)

        ws.cell(dr, 1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False
        )

        for i, c in enumerate(q_cols):
            cell = ws.cell(dr, 2 + i)
            if q_vals[i] is not None:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ltm_cell = ws.cell(dr, ltm_col_idx)
        if ltm_val is not None:
            ltm_cell.number_format = fmt
            ltm_cell.fill = PatternFill("solid", fgColor=CLR_LTM_BG)
        ltm_cell.font      = Font(bold=True, size=10, color=CLR_DARK)
        ltm_cell.alignment = Alignment(horizontal="right", vertical="center")

        if has_qoq:
            cell = ws.cell(dr, qoq_col_idx)
            if qoq_val is not None:
                cell.number_format = PCT_FMT
            cell.font      = Font(size=10, color=CLR_DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        if has_yoy:
            cell = ws.cell(dr, yoy_col_idx)
            if yoy_val is not None:
                cell.number_format = PCT_FMT
            cell.font      = Font(size=10, color=CLR_DARK)
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[dr].height = 15

    # Column widths + freeze
    ws.column_dimensions[get_column_letter(1)].width = 22
    for i in range(len(q_cols)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    ws.column_dimensions[get_column_letter(ltm_col_idx)].width = 13
    if has_qoq:
        ws.column_dimensions[get_column_letter(qoq_col_idx)].width = 9
    if has_yoy:
        ws.column_dimensions[get_column_letter(yoy_col_idx)].width = 9
    ws.freeze_panes = "B3"


# ── Main ──────────────────────────────────────────────────────────────────────
t_start = time.time()
print("[  0.0s] Starting...", flush=True)

t_obj = yf.Ticker(TICKER)
step(f"Fetched ticker info for {TICKER}", t_start)

try:
    info         = t_obj.info
    company_name = info.get("longName", TICKER)
    # financialCurrency = reporting currency (TWD for TSM, EUR for ASML/SPOT)
    # currency          = ADR trading currency (USD for NYSE/NASDAQ-listed ADRs)
    currency     = info.get("financialCurrency") or info.get("currency") or "USD"
except Exception:
    company_name = TICKER
    currency     = "USD"

step("Fetching quarterly data from yfinance...", t_start)
qis = t_obj.quarterly_financials
qbs = t_obj.quarterly_balance_sheet
qcf = t_obj.quarterly_cashflow
step(
    f"Got IS {qis.shape[1]} periods, BS {qbs.shape[1]} periods, CF {qcf.shape[1]} periods",
    t_start,
)

is_cols = _select_cols(qis)
bs_cols = _select_cols(qbs)
cf_cols = _select_cols(qcf)

if not is_cols:
    raise SystemExit(f"No quarterly income statement data available for {TICKER} via yfinance.")

# Console preview — income statement
is_rows = _build_rows(IS_SPEC, qis, is_cols, False)
preview_q = is_cols[-4:]   # show last 4 quarters + LTM in console
hdr = f"{'Label':<22}" + "".join(f"{_q_label(c):>13}" for c in preview_q) + f"{'LTM':>13}"
print(f"\n--- Income Statement ({TICKER}, {currency} Millions, yfinance [restated]) ---", flush=True)
print(hdr, flush=True)
print("-" * len(hdr), flush=True)
for row in is_rows:
    if row["is_header"]:
        print(f"  {row['label'].upper()}", flush=True)
        continue
    fmt_v = (lambda v: f"{v:>12,.2f}") if row["scale"] == 1 else (lambda v: f"{v:>12,.0f}")
    vals  = [fmt_v(row["vals"][c]) if row["vals"].get(c) is not None else f"{'N/A':>12}" for c in preview_q]
    ltm   = row["ltm"]
    ltm_s = fmt_v(ltm) if ltm is not None else f"{'N/A':>12}"
    print(f"  {row['label']:<20}" + "".join(vals) + ltm_s, flush=True)

# Excel export
out_dir  = Path(__file__).resolve().parent.parent / "data"
out_dir.mkdir(exist_ok=True)
out_path = out_dir / f"{TICKER}_quarterly_yf.xlsx"

step(f"\nWriting Excel to {out_path}...", t_start)
wb = Workbook()
wb.remove(wb.active)

STMT_CONFIGS = [
    ("Income Statement", IS_SPEC, qis, is_cols, False),
    ("Balance Sheet",    BS_SPEC, qbs, bs_cols, True),
    ("Cash Flow",        CF_SPEC, qcf, cf_cols, False),
]
for sheet_name, spec, df, cols, is_bs in STMT_CONFIGS:
    write_sheet(wb, spec, df, cols, is_bs, sheet_name, company_name, currency)

if not wb.sheetnames:
    raise SystemExit("No sheets were written.")

wb.save(out_path)
step(f"Done. Open {out_path.name} in Excel - {len(wb.sheetnames)} tab(s).", t_start)
